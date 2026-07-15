"""批量映射：为 清洗完毕_6千组.xlsx 每行的「清洗后名称」找到对应 category_id，
写入新列「对应标准序号」。召回 + DeepSeek 精排，多线程并发。

用法：
    python batch_map.py --limit 20      # 小样测试（每 sheet 前 20 行）
    python batch_map.py                 # 全量 6000 行
"""
import argparse
import json
import os
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl

from product_mapper.agent import ProductMapper
from product_mapper import config

SRC = config.ROOT / "清洗完毕_6千组.xlsx"
PROGRESS = config.ROOT / "cache" / "batch_progress.json"
NAME_COL = "清洗后名称"
OUT_COL = "对应标准序号"


def collect_tasks(wb, limit):
    """返回 [(sheet, row_idx, product)]，row_idx 为 1-based 含表头。"""
    tasks = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        header = [c.value for c in ws[1]]
        name_i = header.index(NAME_COL)
        n = 0
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=name_i + 1).value
            product = (str(val).strip() if val is not None else "")
            tasks.append((sheet, r, product))
            n += 1
            if limit and n >= limit:
                break
    return tasks


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--workers", type=int, default=10)
    args = ap.parse_args()

    print("构建索引…")
    mapper = ProductMapper()
    print(f"索引就绪：{len(mapper.nodes)} 节点，LLM={'开' if config.has_llm() else '关(降级)'}")

    wb = openpyxl.load_workbook(SRC)
    tasks = collect_tasks(wb, args.limit)
    print(f"待映射行数：{len(tasks)}（workers={args.workers}）")

    results = {}   # (sheet,row) -> node_id
    done = 0
    t0 = time.time()

    def work(task):
        sheet, r, product = task
        if not product:
            return sheet, r, None
        try:
            res = mapper.map(product)
            return sheet, r, res.get("node_id")
        except Exception as e:
            return sheet, r, None

    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        futs = [ex.submit(work, t) for t in tasks]
        for f in as_completed(futs):
            sheet, r, nid = f.result()
            results[f"{sheet}|{r}"] = nid
            done += 1
            if done % 50 == 0 or done == len(tasks):
                rate = done / (time.time() - t0)
                eta = (len(tasks) - done) / rate if rate else 0
                print(f"  进度 {done}/{len(tasks)}  {rate:.1f} 行/秒  ETA {eta/60:.1f} 分")
                PROGRESS.parent.mkdir(exist_ok=True)
                PROGRESS.write_text(json.dumps(results, ensure_ascii=False), encoding="utf-8")

    # 写回 xlsx：每个 sheet 追加「对应标准序号」列
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        header = [c.value for c in ws[1]]
        if OUT_COL in header:
            out_i = header.index(OUT_COL) + 1
        else:
            out_i = ws.max_column + 1
            ws.cell(row=1, column=out_i, value=OUT_COL)
        for r in range(2, ws.max_row + 1):
            key = f"{sheet}|{r}"
            if key in results:
                ws.cell(row=r, column=out_i, value=results[key])

    out_path = SRC if not args.limit else config.ROOT / "清洗完毕_6千组_样例.xlsx"
    wb.save(out_path)
    filled = sum(1 for v in results.values() if v is not None)
    print(f"\n完成：写入 {out_path.name}")
    print(f"命中 {filled}/{len(results)} 行（{filled/len(results)*100:.1f}%），"
          f"总耗时 {(time.time()-t0)/60:.1f} 分")


if __name__ == "__main__":
    main()
