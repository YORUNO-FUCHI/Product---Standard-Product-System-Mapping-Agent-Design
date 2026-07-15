"""为 清洗完毕_6千组.xlsx 每行的「清洗后名称」映射 category_id，写入新列「对应标准序号」。

默认【本地模式】：精确匹配 + 融合分 Top-1，不调用 DeepSeek，免费快速。
加 --llm 则启用 DeepSeek 精排（更准但慢、有费用），并发加速。

用法（在 agent/ 目录下）：
    python batch_map_6000.py --limit 20      # 本地模式，小样验证（每 sheet 前 20 行）
    python batch_map_6000.py                 # 本地模式，全量 6000 行
    python batch_map_6000.py --llm --workers 10   # DeepSeek 精排全量
"""
import argparse
import os
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))
os.chdir(str(ROOT))

import openpyxl

from product_mapper import config
from product_mapper.agent import ProductMapper
from product_mapper.pageindex_mapper import PageIndexMapper

SRC = ROOT.parent / "清洗完毕_6千组.xlsx"       # 原文件在上级目录
NAME_COL = "清洗后名称"
OUT_COL = "对应标准序号"


def collect_tasks(wb, limit):
    tasks = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        header = [c.value for c in ws[1]]
        name_i = header.index(NAME_COL)
        n = 0
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=name_i + 1).value
            product = str(val).strip() if val is not None else ""
            tasks.append((sheet, r, product))
            n += 1
            if limit and n >= limit:
                break
    return tasks


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=0, help="每个 sheet 只处理前 N 行（0=全量）")
    ap.add_argument("--llm", action="store_true", help="启用 DeepSeek 精排（默认关闭）")
    ap.add_argument("--route", choices=["a", "b"], default="a",
                    help="a=RAG召回+精排(默认)  b=PageIndex")
    ap.add_argument("--workers", type=int, default=8)
    args = ap.parse_args()

    if not args.llm:
        config.DEEPSEEK_API_KEY = ""   # 本地模式：强制关闭 LLM

    llm_on = args.llm and config.has_llm()
    print(f"路线：Route {args.route.upper()}  |  "
          f"模式：{'DeepSeek' if llm_on else '本地(无DeepSeek)'}")
    print("构建索引…")
    mapper = PageIndexMapper() if args.route == "b" else ProductMapper()
    print(f"索引就绪：{len(mapper.nodes)} 节点")

    if not SRC.exists():
        print(f"找不到源文件：{SRC}")
        return
    wb = openpyxl.load_workbook(SRC)
    tasks = collect_tasks(wb, args.limit)
    print(f"待映射行数：{len(tasks)}，workers={args.workers}")

    results = {}
    t0 = time.time()

    def work(task):
        sheet, r, product = task
        if not product:
            return sheet, r, None
        try:
            return sheet, r, mapper.map(product).get("node_id")
        except Exception:
            return sheet, r, None

    if args.llm:
        with ThreadPoolExecutor(max_workers=args.workers) as ex:
            futs = [ex.submit(work, t) for t in tasks]
            for i, f in enumerate(as_completed(futs), 1):
                sheet, r, nid = f.result()
                results[(sheet, r)] = nid
                if i % 50 == 0 or i == len(tasks):
                    rate = i / (time.time() - t0)
                    print(f"  {i}/{len(tasks)}  {rate:.1f} 行/秒  ETA {(len(tasks)-i)/rate/60:.1f} 分")
    else:
        for i, t in enumerate(tasks, 1):
            sheet, r, nid = work(t)
            results[(sheet, r)] = nid
            if i % 500 == 0 or i == len(tasks):
                rate = i / (time.time() - t0)
                print(f"  {i}/{len(tasks)}  {rate:.1f} 行/秒")

    # 写回：每个 sheet 追加「对应标准序号」列
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        header = [c.value for c in ws[1]]
        out_i = header.index(OUT_COL) + 1 if OUT_COL in header else ws.max_column + 1
        if OUT_COL not in header:
            ws.cell(row=1, column=out_i, value=OUT_COL)
        for r in range(2, ws.max_row + 1):
            if (sheet, r) in results:
                ws.cell(row=r, column=out_i, value=results[(sheet, r)])

    tag = f"route{args.route.upper()}{'_llm' if llm_on else ''}"
    out_path = SRC.parent / (f"清洗完毕_6千组_已映射_{tag}.xlsx" if not args.limit
                             else f"清洗完毕_6千组_样例_{tag}.xlsx")
    wb.save(out_path)
    filled = sum(1 for v in results.values() if v is not None)
    dt = time.time() - t0
    print(f"\n完成 → {out_path}")
    print(f"命中 {filled}/{len(results)}（{filled/len(results)*100:.1f}%），耗时 {dt/60:.2f} 分")


if __name__ == "__main__":
    main()
