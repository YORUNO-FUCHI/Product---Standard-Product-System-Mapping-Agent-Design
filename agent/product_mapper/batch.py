"""Web 批量导入处理：Excel -> A/B/Hybrid 结果 -> Excel 报告。"""
from __future__ import annotations

import time
import re
from pathlib import Path
from typing import Callable

from openpyxl import Workbook, load_workbook

from . import config
from .extension import (
    append_extension_record,
    route_a_reliable,
    route_b_reliable,
    semantic_sim,
)
from .hybrid_decider import choose_hybrid_final


PRODUCT_COLUMN_CANDIDATES = ["清洗后名称", "产品名", "product_name", "name"]
SPLIT_RE = re.compile(r"\s*[、；;]\s*")

DETAIL_HEADERS = [
    "原始行号", "拆分序号", "原始产品串", "拆分后产品名",
    "Route A状态", "Route A是否可靠", "A_node_id", "A_name", "A_path", "A_confidence", "A_source", "A_latency_ms",
    "Route B状态", "Route B是否可靠", "B_node_id", "B_name", "B_path", "B_confidence", "B_source", "B_latency_ms",
    "是否A/B可靠冲突", "是否触发仲裁", "仲裁选择", "仲裁置信度", "仲裁理由", "仲裁是否使用LLM",
    "最终采用路线", "最终node_id", "最终name", "最终path", "最终confidence",
    "是否进入体系扩展", "建议动作", "建议新增节点名", "建议父节点node_id", "建议父节点路径",
    "建议同义词", "建议理由", "优先级", "复核状态", "错误信息",
]

EXTENSION_HEADERS = [
    "拆分后产品名", "Route A状态", "Route B状态", "最接近候选节点", "最接近候选路径",
    "候选分数", "建议动作", "建议新增节点名", "建议父节点node_id", "建议父节点路径",
    "建议同义词", "建议理由", "优先级", "复核状态",
]


class BatchJob:
    def __init__(self, job_id: str, input_path: Path, result_path: Path, mode: str, limit: int):
        self.job_id = job_id
        self.input_path = input_path
        self.result_path = result_path
        self.mode = mode
        self.limit = limit
        self.status = "waiting"
        self.total = 0
        self.processed = 0
        self.stats = {
            "a_hits": 0,
            "b_hits": 0,
            "extensions": 0,
            "errors": 0,
            "conflicts": 0,
            "adjudications": 0,
            "adjudicate_a": 0,
            "adjudicate_b": 0,
            "adjudicate_neither": 0,
        }
        self.preview: list[dict] = []
        self.error = ""
        self.started_at = None
        self.finished_at = None

    def to_dict(self) -> dict:
        return {
            "job_id": self.job_id,
            "status": self.status,
            "mode": self.mode,
            "limit": self.limit,
            "total": self.total,
            "processed": self.processed,
            "stats": self.stats,
            "preview": self.preview,
            "error": self.error,
            "download_ready": self.result_path.exists(),
            "started_at": self.started_at,
            "finished_at": self.finished_at,
        }


def read_products(path: Path, limit: int) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict] = []
    try:
        ws = wb[wb.sheetnames[0]]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return []
        headers = [str(x).strip() if x is not None else "" for x in all_rows[0]]
        col_idx = detect_product_column(headers, all_rows[1:])
        for row_no, values in enumerate(all_rows[1:], start=2):
            value = values[col_idx] if col_idx < len(values) else None
            raw_product = "" if value is None else str(value).strip()
            if not raw_product:
                continue
            for split_seq, product in enumerate(split_products(raw_product), start=1):
                if limit and len(rows) >= limit:
                    return rows
                rows.append({
                    "row_no": row_no,
                    "split_seq": split_seq,
                    "raw_product": raw_product,
                    "product": product,
                })
    finally:
        wb.close()
    return rows


def split_products(text: str) -> list[str]:
    text = (text or "").strip()
    if not text:
        return []
    parts = [part.strip(" \t\r\n,，。") for part in SPLIT_RE.split(text)]
    parts = [part for part in parts if part]
    return parts or [text]


def detect_product_column(headers: list[str], data_rows: list[tuple]) -> int:
    normalized = [h.lower() for h in headers]
    for name in PRODUCT_COLUMN_CANDIDATES:
        if name.lower() in normalized:
            return normalized.index(name.lower())

    best_idx, best_count = 0, -1
    width = max(len(headers), max((len(r) for r in data_rows[:20]), default=0))
    for idx in range(width):
        count = 0
        for row in data_rows[:50]:
            value = row[idx] if idx < len(row) else None
            if isinstance(value, str) and value.strip():
                count += 1
        if count > best_count:
            best_idx, best_count = idx, count
    return best_idx


def process_batch(job: BatchJob, mapper, pageindex_mapper,
                  progress: Callable[[BatchJob], None] | None = None) -> None:
    job.status = "running"
    job.started_at = time.strftime("%Y-%m-%d %H:%M:%S")
    rows = read_products(job.input_path, job.limit)
    job.total = len(rows)
    progress and progress(job)

    original_key = config.DEEPSEEK_API_KEY
    details: list[dict] = []
    extensions: list[dict] = []
    sampled_budget = min(30, max(5, int(job.total * 0.2))) if job.mode == "sampled" else 0
    sampled_used = 0

    try:
        for item in rows:
            product = item["product"]
            row = {
                "原始行号": item["row_no"],
                "拆分序号": item["split_seq"],
                "原始产品串": item["raw_product"],
                "拆分后产品名": product,
                "错误信息": "",
            }
            try:
                use_llm = job.mode == "full"
                if job.mode in {"local", "sampled"}:
                    config.DEEPSEEK_API_KEY = ""
                else:
                    config.DEEPSEEK_API_KEY = original_key

                result_a, _candidates = mapper.explain(product)
                result_b, _trace = pageindex_mapper.explain(product)
                result_b["semantic_sim"] = semantic_sim(mapper, product, result_b)
                a_ok = route_a_reliable(result_a)
                b_ok = route_b_reliable(result_b)

                if job.mode == "sampled" and sampled_used < sampled_budget and should_sample_llm(result_a, result_b, a_ok, b_ok):
                    sampled_used += 1
                    config.DEEPSEEK_API_KEY = original_key
                    result_a, _candidates = mapper.explain(product)
                    result_b, _trace = pageindex_mapper.explain(product)
                    result_b["semantic_sim"] = semantic_sim(mapper, product, result_b)
                    a_ok = route_a_reliable(result_a)
                    b_ok = route_b_reliable(result_b)
                    use_llm = bool(original_key)

                decision = choose_hybrid_final(product, result_a, result_b, mapper, use_llm=use_llm)
                a_ok = decision["a_ok"]
                b_ok = decision["b_ok"]
                extension = decision["extension"]
                final = decision["final"]
                final_route = final.get("route", "")
                adjudication = decision["adjudication"]

                if decision["conflict"]:
                    job.stats["conflicts"] += 1
                if adjudication.get("triggered"):
                    job.stats["adjudications"] += 1
                    choice = adjudication.get("choice")
                    if choice in {"A", "fallback_A"}:
                        job.stats["adjudicate_a"] += 1
                    elif choice == "B":
                        job.stats["adjudicate_b"] += 1
                    elif choice == "neither":
                        job.stats["adjudicate_neither"] += 1

                if extension:
                    append_extension_record(extension)
                    extensions.append(extension)
                    job.stats["extensions"] += 1
                elif final_route in {"Route A", "Hybrid仲裁-Route A", "仲裁失败/降级采用Route A"}:
                    job.stats["a_hits"] += 1
                elif final_route in {"Route B", "Hybrid仲裁-Route B"}:
                    job.stats["b_hits"] += 1

                row.update(flatten_result(result_a, "A", a_ok))
                row.update(flatten_result(result_b, "B", b_ok))
                row.update({
                    "是否A/B可靠冲突": "是" if decision["conflict"] else "否",
                    "是否触发仲裁": "是" if adjudication.get("triggered") else "否",
                    "仲裁选择": adjudication.get("choice", ""),
                    "仲裁置信度": adjudication.get("confidence", ""),
                    "仲裁理由": adjudication.get("reason", ""),
                    "仲裁是否使用LLM": "是" if adjudication.get("llm_used") else "否",
                    "最终采用路线": final_route,
                    "最终node_id": final.get("node_id", ""),
                    "最终name": final.get("name", ""),
                    "最终path": final.get("path", ""),
                    "最终confidence": final.get("confidence", ""),
                    "是否进入体系扩展": "是" if extension else "否",
                    "建议动作": extension.get("action", "") if extension else "",
                    "建议新增节点名": extension.get("new_node_name", "") if extension else "",
                    "建议父节点node_id": extension.get("parent_node_id", "") if extension else "",
                    "建议父节点路径": extension.get("parent_path", "") if extension else "",
                    "建议同义词": "、".join(extension.get("synonyms", [])) if extension else "",
                    "建议理由": extension.get("reason", "") if extension else "",
                    "优先级": extension.get("priority", "") if extension else "",
                    "复核状态": extension.get("review_status", "") if extension else "",
                })
            except Exception as exc:
                job.stats["errors"] += 1
                row.update({"错误信息": str(exc)})
            finally:
                config.DEEPSEEK_API_KEY = original_key

            details.append(row)
            job.processed += 1
            if len(job.preview) < 50:
                job.preview.append(preview_row(row))
            progress and progress(job)

        write_result_workbook(job, details, extensions)
        job.status = "done"
    except Exception as exc:
        job.status = "failed"
        job.error = str(exc)
    finally:
        config.DEEPSEEK_API_KEY = original_key
        job.finished_at = time.strftime("%Y-%m-%d %H:%M:%S")
        progress and progress(job)


def should_sample_llm(result_a: dict, result_b: dict, a_ok: bool, b_ok: bool) -> bool:
    if not a_ok and not b_ok:
        return True
    if result_a.get("node_id") and result_b.get("node_id") and result_a.get("node_id") != result_b.get("node_id"):
        return True
    return float(result_a.get("confidence", 0) or 0) < 0.65 or float(result_b.get("confidence", 0) or 0) < 0.65


def flatten_result(result: dict, prefix: str, reliable: bool) -> dict:
    route = "Route A" if prefix == "A" else "Route B"
    return {
        f"{route}状态": "可靠命中" if reliable else ("弱命中/待复核" if result.get("node_id") else "未命中"),
        f"{route}是否可靠": "是" if reliable else "否",
        f"{prefix}_node_id": result.get("node_id", ""),
        f"{prefix}_name": result.get("name", ""),
        f"{prefix}_path": result.get("path", ""),
        f"{prefix}_confidence": result.get("confidence", 0),
        f"{prefix}_source": result.get("source", ""),
        f"{prefix}_latency_ms": result.get("latency_ms", 0),
    }


def preview_row(row: dict) -> dict:
    return {
        "row_no": row.get("原始行号"),
        "split_seq": row.get("拆分序号"),
        "raw_product": row.get("原始产品串"),
        "product": row.get("拆分后产品名"),
        "route_a": row.get("Route A状态"),
        "route_b": row.get("Route B状态"),
        "final_route": row.get("最终采用路线"),
        "final_node_id": row.get("最终node_id"),
        "final_path": row.get("最终path"),
        "extension": row.get("是否进入体系扩展"),
        "action": row.get("建议动作"),
        "review_status": row.get("复核状态"),
        "error": row.get("错误信息"),
    }


def write_result_workbook(job: BatchJob, details: list[dict], extensions: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "批量汇总"
    ws.append(["指标", "数值"])
    for key, value in [
        ("job_id", job.job_id),
        ("处理模式", job.mode),
        ("总数", job.total),
        ("已处理", job.processed),
        ("Route A可靠命中数", job.stats["a_hits"]),
        ("Route B可靠命中数", job.stats["b_hits"]),
        ("A/B可靠冲突数", job.stats["conflicts"]),
        ("仲裁调用数", job.stats["adjudications"]),
        ("仲裁选择A数量", job.stats["adjudicate_a"]),
        ("仲裁选择B数量", job.stats["adjudicate_b"]),
        ("仲裁判定都不合适数量", job.stats["adjudicate_neither"]),
        ("体系扩展数", job.stats["extensions"]),
        ("错误数", job.stats["errors"]),
        ("开始时间", job.started_at),
        ("结束时间", job.finished_at or ""),
    ]:
        ws.append([key, value])

    ws_detail = wb.create_sheet("批量明细")
    ws_detail.append(DETAIL_HEADERS)
    for row in details:
        ws_detail.append([row.get(h, "") for h in DETAIL_HEADERS])

    ws_ext = wb.create_sheet("体系扩展建议")
    ws_ext.append(EXTENSION_HEADERS)
    for ext in extensions:
        ws_ext.append([
            ext.get("product", ""),
            ext.get("route_a_status", ""),
            ext.get("route_b_status", ""),
            ext.get("nearest_node", ""),
            ext.get("nearest_path", ""),
            ext.get("nearest_score", ""),
            ext.get("action", ""),
            ext.get("new_node_name", ""),
            ext.get("parent_node_id") or "",
            ext.get("parent_path", ""),
            "、".join(ext.get("synonyms") or []),
            ext.get("reason", ""),
            ext.get("priority", ""),
            ext.get("review_status", ""),
        ])

    ws_config = wb.create_sheet("实验配置")
    ws_config.append(["配置项", "值"])
    ws_config.append(["输入文件", str(job.input_path)])
    ws_config.append(["输出文件", str(job.result_path)])
    ws_config.append(["处理模式", job.mode])
    ws_config.append(["最大处理条数", job.limit])
    ws_config.append(["LLM启用", "是" if config.has_llm() else "否"])

    for sheet in wb.worksheets:
        for col in sheet.columns:
            width = max(10, min(60, max(len(str(cell.value or "")) for cell in col) + 2))
            sheet.column_dimensions[col[0].column_letter].width = width
    wb.save(job.result_path)
