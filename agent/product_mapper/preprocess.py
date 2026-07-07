"""数据预处理：清洗公司产品名，去除颜色/尺寸/规格等形容词，提取核心产品名。

输入：../../temp_company_product_0522_1.xlsx（单列 product_name）
输出：cache/cleaned_products.json / cache/cleaned_products.xlsx

清洗策略（按顺序执行）：
  1. 去括号内容（含中英文括号、全半角）
  2. 去掉前导数字+单位组合（如 "12英寸"、"1200V"、"280Ah"）
  3. 去掉颜色词（红/黑/白/蓝...）
  4. 去掉规格/品质形容词（高端/智能/原装/大型...）
  5. 去掉残留的纯规格型号行（清洗后为空或纯数字的标记为待审核）
  6. 后处理：过度清洗保护（结果过短时回退到保守清洗）
"""

import json
import re
import os
from pathlib import Path

# ── 路径 ────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent
EXCEL_IN = ROOT.parent / "temp_company_product_0522_1.xlsx"
CACHE_DIR = ROOT / "cache"
CACHE_DIR.mkdir(exist_ok=True)
JSON_OUT = CACHE_DIR / "cleaned_products.json"
EXCEL_OUT = CACHE_DIR / "cleaned_products.xlsx"

# ── 颜色词（中文 + 英文）────────────────────────────────────────
COLORS = [
    "红色", "橙色", "黄色", "绿色", "青色", "蓝色", "紫色",
    "黑色", "白色", "灰色", "金色", "银色", "粉色", "棕色", "褐色",
    "透明", "无色", "原色", "彩色", "混色",
    # 英文
    "red", "orange", "yellow", "green", "cyan", "blue", "purple",
    "black", "white", "gray", "grey", "gold", "silver", "pink", "brown",
    # 缩写/前缀
    "深", "浅", "暗", "亮", "淡", "艳",
]

# ── 规格单位模式（长的在前，避免 m3 被 m 误匹配）────────────────
_UNITS = [
    # 体积/容量（长的在前）
    "立方[米厘米]", "m3", "cm3",
    "毫升", "加仑", "ml", "gal",
    # 长度
    "英寸/秒", "ips",
    "英寸", "英尺", "厘米", "毫米", "微米", "纳米",
    "吋", "寸", "米", "mm", "cm", "dm", "μm", "nm", "km",
    # 数据量/带宽（在 g/m 之前，避免 Gb 被 g 误匹配）
    "[KMGTP]bps", "[KMGTP]?bps", "Gbps", "Mbps", "kbps",
    "[KMGTP]?Bps",
    "Gb", "Mb", "Kb", "Tb", "Pb",
    "GB", "MB", "KB", "TB", "PB",
    # 重量
    "千克", "公斤", "毫克", "mg", "kg", "克", "吨", "吨级", "斤", "两",
    "t", "lb", "oz", "g",
    # 电压
    "千伏", "毫伏", "kV", "mV", "MV", "伏", "V",
    # 电流
    "毫安", "安时", "毫安时", "Ah", "mAh", "mA", "安", "A",
    # 功率/能量
    "千瓦时", "兆瓦", "毫瓦", "瓦时",
    "kWp", "Wp",  # 太阳能峰值功率
    "kW", "MW", "mW", "Wh", "kWh", "千瓦", "瓦", "W",
    # 频率
    "千赫", "兆赫", "GHz", "MHz", "kHz",
    "Hz",
    # 像素
    "万像素", "像素", "MP",
    # 中文规格
    "位", "级", "座", "轴", "缸", "芯",
    # 新增：微米变体 / 载重吨 / 分辨率K
    "μm", "um",  # 微米
    "DWT",      # 载重吨
    "dwt",
    "K",         # 分辨率（4K, 8K）
    # 长度/距离
    "米", "m",  # m 放最后，避免误匹配含 m 的其他单位
    "升", "L",
]
_UNIT_STR = "|".join(_UNITS)
UNIT_PATTERN = re.compile(
    r"[\d.]+[/\d.]*\s*"  # 数字部分（含小数点、斜杠）
    r"(?:" + _UNIT_STR + r")",
    re.IGNORECASE,
)

# ── 品质/规格形容词 ─────────────────────────────────────────────
QUALITY_WORDS = [
    "高端", "中端", "低端", "旗舰", "入门", "顶配",
    "智能", "智慧", "自动", "手动", "气动", "液压",
    "原装", "进口", "国产", "代工", "定制",
    "大型", "中型", "小型", "微型", "迷你", "便携", "手持",
    "超薄", "超轻", "超重", "加厚", "加长",
    "高速", "低速", "快速", "慢速", "极速",
    "高精度", "高灵敏", "高功率", "高效率", "高性能",
    "低功耗", "低噪声", "低损耗",
    "防水", "防尘", "防爆", "防腐", "耐高温", "耐低温", "耐压",
    "环保", "绿色", "节能", "减排",
    "全新", "全新原装", "原厂原装",
    "二手", "翻新", "库存", "尾货",
    "测试", "实验", "科研", "工业级", "民用", "军用",
    "超大型", "超大", "重型", "轻型",
]

# ── 可配置停用词（先留空，后续按需添加）───────────────────────
STOP_WORDS = []


# ── 化学手性标记（保留，不去除）────────────────────────────────
# (R)/(S) 手性、(E)/(Z) 顺反异构、(D)/(L) 旋光、(±)/(+)/(-) 外消旋等
_STEREO_PREFIX = re.compile(
    r"[（(]"
    r"(?:[RESZDL]|[RS][ESZ]?|DL|[±+\-]|"
    r"(?:R,R|S,S|R,S|S,R|E,E|Z,Z|E,Z|Z,E)|"  # 多手性中心
    r"[A-Za-z]\d?"  # 如 H1, C2 等位置标记
    r")"
    r"[）)]\s*[-—]?"  # 后跟可选连字符
)


def remove_brackets(text: str) -> str:
    """去除括号及内容，但保留化学手性标记如 (R)-, (S)-, (E)-, (±)- 等。"""
    # 先保护化学前缀
    protected = []
    def _protect(m):
        protected.append(m.group(0))
        return f"《CHEM{len(protected) - 1}》"

    text = _STEREO_PREFIX.sub(_protect, text)

    # 去除其余括号及内容（中英文、全半角）
    text = re.sub(r"[（(][^）)]*[）)]", "", text)
    text = re.sub(r"\([^)]*\)", "", text)
    text = re.sub(r"\[[^\]]*\]", "", text)

    # 还原化学前缀（必须在书名号清理之前，否则 《CHEM0》 会被误删）
    for i, s in enumerate(protected):
        text = text.replace(f"《CHEM{i}》", s)

    # 书名号、尖括号
    text = re.sub(r"《[^》]*》", "", text)
    text = re.sub(r"<[^>]*>", "", text)

    return text


def remove_leading_unit(text: str) -> str:
    """去掉前导的数字+单位组合，如 '12英寸CMP设备' → 'CMP设备'。"""
    prev = None
    max_iter = 10
    while prev != text and max_iter > 0:
        max_iter -= 1
        prev = text
        text = UNIT_PATTERN.sub("", text, count=1).strip()
        # 清理残留：前导数字+空格（但不吃紧接大写字母的，如 0BB 技术名）
        text = re.sub(r"^[\d.]+[/\d.\-xX×]*[\d.]*\s+(?=[一-鿿A-Za-z])", "", text).strip()
        # 清理残留：单位被匹配后留下的纯数字碎片
        text = re.sub(r"^[\d.\-]+(?:[/xX×][\d.]+)?\s+(?=[一-鿿])", "", text).strip()
        # 新增：清理 "XX.XMW"、"XXkA" 这类粘连单位
        text = re.sub(r"^[\d.]+[/\d.\-xX×]*\s*[A-Za-z]+\s+(?=[一-鿿])", "", text).strip()
    return text


def remove_colors(text: str) -> str:
    """去掉颜色形容词。"""
    for c in COLORS:
        # 匹配词首或词中的颜色词（前可接数字，后可接 '色'）
        text = re.sub(rf"\d*{c}色?", "", text, flags=re.IGNORECASE)
    return text


def remove_quality(text: str) -> str:
    """去掉品质/规格形容词。"""
    for w in QUALITY_WORDS:
        text = text.replace(w, "")
    return text


def remove_leading_model(text: str) -> str:
    """去掉前导的纯型号/编码/品牌名（如 'AK3221M', 'I-FLASH', 'BLUETTI'）。

    改进：检查去除后的内容长度，如果过短则不执行去除。
    """
    # 完整型号前缀（可能紧跟中文，无空格）
    m = re.match(
        r"^([A-Za-z]"                               # 字母开头
        r"(?:[-A-Za-z0-9._/+]{0,29}"                 # 字母数字特殊符号
        r"|[A-Za-z0-9][-A-Za-z0-9._/+]{0,28})"       # 数字+字母开头
        r"|"                                         # 或
        r"[A-Za-z][-A-Za-z0-9._/+]{1,15}(?:\s+[A-Za-z][-A-Za-z0-9._/+]{1,15}){0,2}"  # 多段
        r")"
        r"(?:\s+|$|(?=[一-鿿]))"  # 后跟空格、结束、或直接粘中文
        r"(.+)$",
        text)
    if m:
        prefix, rest = m.group(1), m.group(2)
        # 只当前缀不含中文，且剩余内容非空，且剩余长度>=3或占比>=25%时才去掉
        if not re.search(r"[一-鿿]", prefix) and rest and len(rest.strip()) > 0:
            rest_clean = rest.strip()
            if len(rest_clean) >= 3 or len(rest_clean) >= len(text) * 0.25:
                return rest_clean
    return text


def remove_slash_alternatives(text: str) -> str:
    """处理斜杠分隔的多方案表达，如 '存储类(SRAM/DRAM/NOR Flash)' → 保留主体。

    策略：去掉括号内纯由斜杠分隔的英文/数字组合，其余保留。
    """
    # 中文括号内的斜杠列表：如（SRAM/DRAM/NOR Flash）
    text = re.sub(
        r"[（(]\s*[A-Za-z0-9.\-]+(?:/[A-Za-z0-9.\-\s]+)+\s*[）)]",
        "", text
    )
    return text


def clean_product(raw: str) -> dict:
    """清洗单条产品名，返回 {raw, cleaned, status}。"""
    if not raw or not isinstance(raw, str):
        return {"raw": str(raw or ""), "cleaned": "", "status": "empty"}

    original = raw.strip()
    text = original

    # 1) 去括号
    text = remove_brackets(text).strip()

    # 1.5) 去斜杠多方案括号
    text = remove_slash_alternatives(text).strip()

    # 2) 去前导规格型号 + 数字单位
    text = remove_leading_unit(text).strip()

    # 3) 去前导纯型号编码
    text = remove_leading_model(text).strip()

    # 4) 去颜色词
    text = remove_colors(text).strip()

    # 5) 去品质形容词
    text = remove_quality(text).strip()

    # 6) 清理多余空白和残留标点
    text = re.sub(r"\s{2,}", " ", text)
    # 去掉范围残留（如 "100-"、"150V→150" 后的 "-"）
    text = re.sub(r"^[\d.\-]+[\-—]\s*", "", text)
    text = re.sub(r"^[,，.。、/\\|;；:：\-—_\s]+", "", text)
    text = re.sub(r"[,，.。、/\\|;；:：\-—_\s]+$", "", text)
    text = text.strip()

    # ── 后处理：过度清洗保护 ──
    has_chinese = bool(re.search(r"[一-鿿]", original))

    # 若原文本有中文，但清洗结果过短（<=2字符 或 长度占比<20%），
    # 回退到保守清洗（仅去括号+去前导单位，保留型号和品质词）
    if has_chinese and text and len(text) <= 2:
        # 保守清洗：仅去括号 + 去前导单位
        conservative = remove_brackets(original).strip()
        conservative = remove_leading_unit(conservative).strip()
        conservative = re.sub(r"\s{2,}", " ", conservative)
        conservative = re.sub(r"^[,，.。、/\\|;；:：\-—_\s]+", "", conservative)
        conservative = re.sub(r"[,，.。、/\\|;；:：\-—_\s]+$", "", conservative)
        conservative = conservative.strip()
        if len(conservative) > len(text) and len(conservative) >= 3:
            text = conservative

    # 若原文本有中文，清洗结果占比过低（<20%），回退
    if has_chinese and text and len(original) > 8:
        ratio = len(text) / len(original)
        if ratio < 0.20:
            conservative = remove_brackets(original).strip()
            conservative = remove_leading_unit(conservative).strip()
            conservative = re.sub(r"\s{2,}", " ", conservative)
            conservative = re.sub(r"^[,，.。、/\\|;；:：\-—_\s]+", "", conservative)
            conservative = re.sub(r"[,，.。、/\\|;；:：\-—_\s]+$", "", conservative)
            conservative = conservative.strip()
            if len(conservative) > len(text):
                text = conservative

    # ── 判断状态 ──
    if not text or len(text) <= 1:
        return {"raw": original, "cleaned": text, "status": "empty_after_clean"}

    # 清洗后只剩纯数字/字母/符号（无中文），可能是规格型号
    if not re.search(r"[一-鿿]", text):
        # 但若原始有中文且清洗后有含义的英文缩写（>=3字符），可能是产品名
        if has_chinese and len(text) >= 3:
            return {"raw": original, "cleaned": text, "status": "cleaned"}
        if len(text) <= 20:
            return {"raw": original, "cleaned": text, "status": "likely_spec"}

    if text == original:
        return {"raw": original, "cleaned": text, "status": "unchanged"}

    return {"raw": original, "cleaned": text, "status": "cleaned"}


def main(limit: int = None, show_samples: int = 30):
    """读取 Excel → 逐行清洗 → 输出 JSON + Excel。"""
    import openpyxl

    print(f"读取: {EXCEL_IN}")
    wb = openpyxl.load_workbook(EXCEL_IN, read_only=True, data_only=True)
    ws = wb["temp_company_product_0522_1"]
    rows = list(ws.iter_rows(values_only=True))
    header = str(rows[0][0]) if rows[0][0] else "product_name"
    data = [str(r[0]).strip() for r in rows[1:] if r[0]]
    total = len(data)
    print(f"总行数: {total}")

    if limit:
        data = data[:limit]
        print(f"限制: {limit} 条")

    results = []
    stats = {"cleaned": 0, "unchanged": 0, "empty_after_clean": 0, "likely_spec": 0, "empty": 0}

    for i, raw in enumerate(data):
        r = clean_product(raw)
        stats[r["status"]] = stats.get(r["status"], 0) + 1
        results.append(r)
        if (i + 1) % 100000 == 0:
            print(f"  已处理 {i + 1}/{len(data)} ...")

    print(f"\n=== 清洗统计 ===")
    print(f"总记录:        {len(results)}")
    print(f"已清洗:        {stats.get('cleaned', 0)}")
    print(f"无变化:        {stats.get('unchanged', 0)}")
    print(f"清洗后为空:    {stats.get('empty_after_clean', 0)}")
    print(f"疑似规格型号:  {stats.get('likely_spec', 0)}")

    # ── 输出 JSON ──
    JSON_OUT.write_text(json.dumps(results, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"\nJSON 输出: {JSON_OUT}")

    # ── 输出 Excel ──
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "cleaned"
    ws_out.append(["序号", "原始名称", "清洗后名称", "状态"])
    for i, r in enumerate(results, 1):
        ws_out.append([i, r["raw"], r["cleaned"], r["status"]])
    wb_out.save(EXCEL_OUT)
    print(f"Excel 输出: {EXCEL_OUT}")

    # ── 展示样例 ──
    if show_samples:
        cleaned_samples = [r for r in results if r["status"] == "cleaned"]
        import random
        random.seed(1)
        samples = random.sample(cleaned_samples, min(show_samples, len(cleaned_samples)))
        print(f"\n=== 清洗样例（{len(samples)} 条）===")
        for s in samples:
            print(f"  [{s['status']}] {s['raw'][:60]}")
            print(f"           → {s['cleaned'][:60]}")
            print()


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="清洗公司产品名")
    parser.add_argument("--limit", type=int, default=None, help="限制处理行数")
    parser.add_argument("--samples", type=int, default=30, help="展示样例数")
    args = parser.parse_args()
    main(limit=args.limit, show_samples=args.samples)
