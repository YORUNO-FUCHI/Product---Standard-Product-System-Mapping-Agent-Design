"""体系扩展建议：双路线未可靠命中后的可复核建议流程。"""
from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from . import config
from .agent import ProductMapper
from .llm import chat_json
from .rerank import _fuse


EXTENSION_JSONL = config.CACHE_DIR / "extension_suggestions.jsonl"
EXTENSION_XLSX = config.CACHE_DIR / "extension_suggestions.xlsx"

EXTENSION_HEADERS = [
    "时间", "产品名", "Route A状态", "Route B状态", "最接近候选节点", "最接近候选路径",
    "候选分数", "建议动作", "建议新增节点名", "建议父节点node_id", "建议父节点路径",
    "建议同义词", "建议理由", "优先级", "复核状态",
]

EXTENSION_ACTIONS = {
    "清洗异常/不扩展",
    "补同义词",
    "新增叶子节点",
    "新增中间节点",
    "体系外/新增大类",
}

EXTENSION_SYSTEM = """你是产品标准体系维护助手。请判断一个未命中产品应该如何处理。
只能从这些建议动作中选择一个：清洗异常/不扩展、补同义词、新增叶子节点、新增中间节点、体系外/新增大类。
严格输出 JSON：
{"action":"<建议动作>","new_node_name":"<建议新增节点名或空>","parent_node_id":<int或null>,"parent_path":"<建议父节点路径或空>","synonyms":["<建议同义词>"],"reason":"<简短理由>","priority":"高/中/低"}"""


def looks_damaged(product: str) -> bool:
    product = (product or "").strip()
    if len(product) <= 2:
        return True
    if re.fullmatch(r"[-A-Za-z0-9._/+\\ ]{1,20}", product):
        return True
    return "?" in product or "锟" in product


def hit(result: dict | None) -> bool:
    return bool(result and result.get("node_id") is not None)


def route_a_reliable(result: dict | None) -> bool:
    if not hit(result):
        return False
    source = result.get("source", "")
    confidence = float(result.get("confidence", 0.0) or 0.0)
    if source in {"exact_match", "hybrid_raga"} and confidence >= 0.85:
        return True
    if source == "llm" and confidence >= 0.70:
        return True
    return source == "fusion" and confidence >= 0.85


def route_b_reliable(result: dict | None) -> bool:
    if not hit(result):
        return False
    source = result.get("source", "")
    confidence = float(result.get("confidence", 0.0) or 0.0)
    if source in {"pageindex_exact", "hybrid_pageindex"} and confidence >= 0.85:
        return True
    if source == "pageindex" and confidence >= 0.70:
        return True
    return source == "pageindex_trigram" and confidence >= 0.18


def status_label(result: dict | None, reliable: bool) -> str:
    if not hit(result):
        return "未命中"
    if reliable:
        return "可靠命中"
    return "弱命中/待复核"


def nearest_taxonomy_candidates(mapper: ProductMapper, product: str, top_k: int = 5) -> list[dict]:
    original_key = config.DEEPSEEK_API_KEY
    config.DEEPSEEK_API_KEY = ""
    try:
        ordered = _fuse(mapper.recaller.recall(product))[:top_k]
    finally:
        config.DEEPSEEK_API_KEY = original_key

    items = []
    for cand in ordered:
        node = cand.node
        if node.is_leaf and len(node.path_names) > 1:
            parent_id = node.parent_id
            parent_path = " > ".join(node.path_names[:-1])
        else:
            parent_id = node.id
            parent_path = node.path_str
        items.append({
            "node_id": node.id,
            "name": node.name,
            "path": node.path_str,
            "parent_id": parent_id,
            "parent_path": parent_path,
            "is_leaf": node.is_leaf,
            "score": round(float(getattr(cand, "fused", 0.0)), 3),
            "trgm": round(float(getattr(cand, "trgm", 0.0)), 3),
            "vec": round(float(getattr(cand, "vec", 0.0)), 3),
        })
    return items


def _format_candidates(candidates: list[dict]) -> str:
    if not candidates:
        return "无候选"
    return "\n".join(
        f"{i}. node_id={c['node_id']} | {c['path']} | score={c['score']}, trgm={c['trgm']}, vec={c['vec']}"
        for i, c in enumerate(candidates, 1)
    )


def _normalize_llm_output(out: dict | None, product: str, candidates: list[dict]) -> dict:
    best = candidates[0] if candidates else {}
    if not out:
        if looks_damaged(product):
            action = "清洗异常/不扩展"
            reason = "产品名过短、疑似型号或含异常字符，先复核清洗结果"
            priority = "低"
        elif best and best.get("score", 0.0) >= 1.0:
            action = "补同义词"
            reason = "存在相近标准节点，优先作为同义词候选人工确认"
            priority = "中"
        elif best and best.get("score", 0.0) >= 0.45:
            action = "新增叶子节点"
            reason = "存在相近父路径，但当前体系缺少更细产品节点"
            priority = "中"
        else:
            action = "体系外/新增大类"
            reason = "没有足够相近的候选路径，建议人工判断是否属于体系外领域"
            priority = "低"
        return {
            "action": action,
            "new_node_name": "" if action in {"清洗异常/不扩展", "补同义词"} else product,
            "parent_node_id": best.get("parent_id"),
            "parent_path": best.get("parent_path", ""),
            "synonyms": [product] if action == "补同义词" else [],
            "reason": reason,
            "priority": priority,
        }

    action = str(out.get("action") or "").strip()
    if action not in EXTENSION_ACTIONS:
        action = "体系外/新增大类"
    synonyms = out.get("synonyms") or []
    if not isinstance(synonyms, list):
        synonyms = [str(synonyms)]
    parent_node_id = out.get("parent_node_id")
    try:
        parent_node_id = int(parent_node_id) if parent_node_id not in (None, "") else None
    except Exception:
        parent_node_id = None
    priority = str(out.get("priority") or "中").strip()
    if priority not in {"高", "中", "低"}:
        priority = "中"
    return {
        "action": action,
        "new_node_name": str(out.get("new_node_name") or "").strip(),
        "parent_node_id": parent_node_id,
        "parent_path": str(out.get("parent_path") or "").strip(),
        "synonyms": [str(x).strip() for x in synonyms if str(x).strip()],
        "reason": str(out.get("reason") or "").strip(),
        "priority": priority,
    }


def suggest_extension(product: str, mapper: ProductMapper, route_a: dict | None = None,
                      route_b: dict | None = None, use_llm: bool = True) -> dict:
    candidates = nearest_taxonomy_candidates(mapper, product, top_k=5)
    out = None
    if use_llm and config.has_llm():
        user = (
            f"未命中产品名：{product}\n"
            f"Route A状态：{status_label(route_a, route_a_reliable(route_a))}\n"
            f"Route B状态：{status_label(route_b, route_b_reliable(route_b))}\n\n"
            f"最接近候选节点 Top5：\n{_format_candidates(candidates)}"
        )
        out = chat_json(EXTENSION_SYSTEM, user, timeout=60)
    suggestion = _normalize_llm_output(out, product, candidates)
    best = candidates[0] if candidates else {}
    return {
        "product": product,
        "route_a_status": status_label(route_a, route_a_reliable(route_a)),
        "route_b_status": status_label(route_b, route_b_reliable(route_b)),
        "nearest_node": best.get("name", ""),
        "nearest_path": best.get("path", ""),
        "nearest_score": best.get("score", ""),
        "action": suggestion["action"],
        "new_node_name": suggestion["new_node_name"],
        "parent_node_id": suggestion["parent_node_id"],
        "parent_path": suggestion["parent_path"],
        "synonyms": suggestion["synonyms"],
        "reason": suggestion["reason"],
        "priority": suggestion["priority"],
        "review_status": "待复核",
        "candidates": candidates,
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }


def append_extension_record(extension: dict, jsonl_path: Path = EXTENSION_JSONL,
                            xlsx_path: Path = EXTENSION_XLSX) -> None:
    config.CACHE_DIR.mkdir(exist_ok=True)
    with jsonl_path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(extension, ensure_ascii=False, default=str) + "\n")

    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "体系扩展建议"
        ws.append(EXTENSION_HEADERS)
    ws.append(flatten_extension_record(extension))
    wb.save(xlsx_path)


def flatten_extension_record(extension: dict) -> list[Any]:
    return [
        extension.get("created_at", ""),
        extension.get("product", ""),
        extension.get("route_a_status", ""),
        extension.get("route_b_status", ""),
        extension.get("nearest_node", ""),
        extension.get("nearest_path", ""),
        extension.get("nearest_score", ""),
        extension.get("action", ""),
        extension.get("new_node_name", ""),
        extension.get("parent_node_id") or "",
        extension.get("parent_path", ""),
        "、".join(extension.get("synonyms") or []),
        extension.get("reason", ""),
        extension.get("priority", ""),
        extension.get("review_status", ""),
    ]
