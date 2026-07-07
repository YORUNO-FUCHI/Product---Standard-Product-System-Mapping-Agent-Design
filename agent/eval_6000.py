"""评测清洗后 6000 条数据：Route A / Route B / Hybrid 对比"""
import json, time, sys, os
from pathlib import Path
from collections import Counter

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

os.chdir(str(ROOT))

from product_mapper import config
from product_mapper.taxonomy import load_nodes
from product_mapper.agent import ProductMapper
from product_mapper.pageindex_mapper import PageIndexMapper

XLSX = ROOT.parent / "清洗完毕_6千组.xlsx"
print(f"XLSX path: {XLSX}, exists={XLSX.exists()}")

def load_6000():
    import openpyxl
    wb = openpyxl.load_workbook(str(XLSX), read_only=True, data_only=True)
    records = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        for r in rows[1:]:
            raw = str(r[1]).strip() if r[1] else ""
            cleaned = str(r[2]).strip() if r[2] else ""
            records.append({"raw": raw, "cleaned": cleaned})
    wb.close()
    return records

def run_eval(records, mapper_a, mapper_b):
    results = []
    n = len(records)
    for i, rec in enumerate(records):
        product = rec["cleaned"] if rec["cleaned"] else rec["raw"]
        if not product:
            results.append(None)
            continue

        r = {}

        # Route A
        t0 = time.time()
        res_a = mapper_a.map(product)
        r["a_hit"] = res_a["node_id"] is not None
        r["a_conf"] = round(res_a.get("confidence", 0), 3)
        r["a_source"] = res_a.get("source", "")
        r["a_node"] = res_a.get("node_id")
        r["a_ms"] = round((time.time() - t0) * 1000, 1)

        # Route B
        t0 = time.time()
        res_b = mapper_b.map(product)
        r["b_hit"] = res_b["node_id"] is not None
        r["b_conf"] = round(res_b.get("confidence", 0), 3)
        r["b_source"] = res_b.get("source", "")
        r["b_node"] = res_b.get("node_id")
        r["b_layers"] = res_b.get("n_layers_visited", 0)
        r["b_ms"] = round((time.time() - t0) * 1000, 1)

        # Hybrid
        if r["a_hit"] and (r["a_conf"] >= 0.85 or r["a_source"] == "exact_match"
                           or (r["a_source"] == "llm" and r["a_conf"] >= 0.7)):
            r["h_source"] = "hybrid_raga"
            r["h_hit"] = r["a_hit"]
            r["h_conf"] = r["a_conf"]
            r["h_node"] = r["a_node"]
        elif r["b_hit"] and (not r["a_hit"] or r["b_conf"] > r["a_conf"]):
            r["h_source"] = "hybrid_pageindex"
            r["h_hit"] = r["b_hit"]
            r["h_conf"] = r["b_conf"]
            r["h_node"] = r["b_node"]
        else:
            r["h_source"] = "hybrid_raga" if r["a_hit"] else "hybrid_empty"
            r["h_hit"] = r["a_hit"]
            r["h_conf"] = r["a_conf"]
            r["h_node"] = r["a_node"]

        results.append(r)

        if (i + 1) % 500 == 0:
            print(f"  {i+1}/{n}")

    return results

def compute_stats(results, label):
    valid = [r for r in results if r is not None]
    n = len(valid)
    if n == 0:
        return {}

    s = {}
    s["n"] = n
    s["label"] = label

    # Route A
    s["a_hit"] = sum(1 for r in valid if r["a_hit"])
    s["a_hit_rate"] = s["a_hit"] / n
    s["a_avg_conf"] = sum(r["a_conf"] for r in valid) / n
    s["a_avg_ms"] = sum(r["a_ms"] for r in valid) / n
    s["a_sources"] = dict(Counter(r["a_source"] for r in valid).most_common())

    # Route B
    s["b_hit"] = sum(1 for r in valid if r["b_hit"])
    s["b_hit_rate"] = s["b_hit"] / n
    s["b_avg_conf"] = sum(r["b_conf"] for r in valid) / n
    s["b_avg_ms"] = sum(r["b_ms"] for r in valid) / n
    s["b_avg_layers"] = sum(r["b_layers"] for r in valid) / n if valid else 0
    s["b_sources"] = dict(Counter(r["b_source"] for r in valid).most_common())

    # Hybrid
    s["h_hit"] = sum(1 for r in valid if r["h_hit"])
    s["h_hit_rate"] = s["h_hit"] / n
    s["h_avg_conf"] = sum(r["h_conf"] for r in valid) / n
    s["h_sources"] = dict(Counter(r["h_source"] for r in valid).most_common())

    # Agreement
    ab_both = [r for r in valid if r["a_hit"] and r["b_hit"]]
    s["agree_ab"] = sum(1 for r in ab_both if r["a_node"] == r["b_node"]) / max(1, len(ab_both))
    ah_both = [r for r in valid if r["a_hit"] and r["h_hit"]]
    s["agree_ah"] = sum(1 for r in ah_both if r["a_node"] == r["h_node"]) / max(1, len(ah_both))
    all_hit = [r for r in valid if r["a_hit"] and r["b_hit"] and r["h_hit"]]
    s["all_agree"] = sum(1 for r in all_hit if r["a_node"] == r["b_node"] == r["h_node"]) / max(1, len(all_hit))

    # Confidence buckets
    buckets = [(0, 0.3), (0.3, 0.5), (0.5, 0.7), (0.7, 0.85), (0.85, 1.01)]
    s["conf_dist"] = {}
    for lo, hi in buckets:
        cnt = sum(1 for r in valid if lo <= r["h_conf"] < hi)
        s["conf_dist"][f"{lo}-{hi}"] = cnt / n

    return s

def print_stats(s):
    print(f"\n{'='*60}")
    print(f"  {s['label']} ({s['n']} 条)")
    print(f"{'='*60}")
    print(f"  {'指标':<25} {'Route A':<15} {'Route B':<15} {'Hybrid':<15}")
    print(f"  {'-'*25} {'-'*15} {'-'*15} {'-'*15}")
    print(f"  {'命中率':<25} {s['a_hit_rate']:<14.1%} {s['b_hit_rate']:<14.1%} {s['h_hit_rate']:<14.1%}")
    print(f"  {'命中数':<25} {s['a_hit']:<14} {s['b_hit']:<14} {s['h_hit']:<14}")
    print(f"  {'平均置信度':<25} {s['a_avg_conf']:<14.3f} {s['b_avg_conf']:<14.3f} {s['h_avg_conf']:<14.3f}")
    print(f"  {'平均延迟(ms)':<25} {s['a_avg_ms']:<14.1f} {s['b_avg_ms']:<14.1f} {'--':<15}")
    print(f"  {'B 平均搜索层数':<25} {'--':<15} {s['b_avg_layers']:<14.1f} {'--':<15}")
    print(f"\n  来源分布:")
    print(f"    Route A: {s['a_sources']}")
    print(f"    Route B: {s['b_sources']}")
    print(f"    Hybrid:  {s['h_sources']}")
    print(f"\n  一致性(same node_id):")
    print(f"    A-B: {s['agree_ab']:.1%}  A-H: {s['agree_ah']:.1%}  三者一致: {s['all_agree']:.1%}")
    print(f"\n  Hybrid 置信度分布:")
    for k, v in s['conf_dist'].items():
        bar = '#' * int(v * 50)
        print(f"    {k}: {v:.1%} {bar}")


if __name__ == "__main__":
    records = load_6000()
    print(f"加载清洗数据: {len(records)} 条")
    print(f"LLM 状态: {'已启用' if config.has_llm() else '未启用'}")

    print("\n>>> 初始化...")
    nodes = load_nodes()
    mapper_a = ProductMapper(nodes)
    mapper_b = PageIndexMapper(nodes)
    print(f"节点数: {len(nodes)}")

    has_api = config.has_llm()

    # ── 基线：LLM=OFF 全量 ──
    print(f"\n>>> 基线测试：6000 条全量 (LLM=OFF)")
    config.DEEPSEEK_API_KEY = ""
    t0 = time.time()
    results_base = run_eval(records, mapper_a, mapper_b)
    dt = time.time() - t0
    print(f"  总耗时: {dt:.1f}s ({dt/len(records)*1000:.1f}ms/条)")
    stats_base = compute_stats(results_base, "6000条 基线 (LLM=OFF)")
    print_stats(stats_base)

    # ── LLM 样本 ──
    if has_api:
        os.environ["DEEPSEEK_API_KEY"] = os.environ.get("DEEPSEEK_API_KEY_BAK", "")
        # re-read from .env
        from dotenv import load_dotenv
        load_dotenv(ROOT / ".env")
        import importlib
        importlib.reload(config)

        if config.has_llm():
            n_sample = 200
            print(f"\n>>> LLM 样本测试：{n_sample} 条 (LLM=ON)")
            t0 = time.time()
            results_llm = run_eval(records[:n_sample], mapper_a, mapper_b)
            dt = time.time() - t0
            print(f"  总耗时: {dt:.1f}s ({dt/n_sample*1000:.1f}ms/条)")
            stats_llm = compute_stats(results_llm, f"{n_sample}条 样本 (LLM=ON)")
            print_stats(stats_llm)
        else:
            print("\nLLM 重新加载失败，跳过 LLM 样本测试")
            has_api = False

    # ── 保存 ──
    out = {"baseline_6000_no_llm": {k: v for k, v in stats_base.items() if k != "label"}}
    if has_api:
        out["sample_200_with_llm"] = {k: v for k, v in stats_llm.items() if k != "label"}

    out_path = ROOT / "cache" / "eval_6000_results.json"
    out_path.write_text(json.dumps(out, ensure_ascii=False, indent=1, default=str), encoding="utf-8")
    print(f"\n结果已保存: {out_path}")
