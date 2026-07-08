"""Low-token Route A / Route B evaluation for cleaned product names.

The script intentionally avoids true accuracy metrics because there is no
reliable gold-label workbook yet. It measures coverage, confidence, consistency,
conflicts, review load, and sampled LLM cost instead.
"""
from __future__ import annotations

import argparse
import json
import random
import re
import sys
import time
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
PROJECT_ROOT = ROOT.parent
sys.path.insert(0, str(ROOT))

from product_mapper import config  # noqa: E402
from product_mapper.agent import ProductMapper  # noqa: E402
from product_mapper.embedder import st_available  # noqa: E402
from product_mapper.llm import chat_json, get_llm_stats, reset_llm_stats  # noqa: E402
from product_mapper.pageindex_mapper import PageIndexMapper  # noqa: E402
from product_mapper.rerank import _fuse  # noqa: E402
from product_mapper.taxonomy import load_nodes  # noqa: E402


INPUT_XLSX = PROJECT_ROOT / "清洗完毕_6千组.xlsx"
OUTPUT_XLSX = ROOT / "cache" / "route_ab_balanced_results.xlsx"
JSONL_PATH = ROOT / "cache" / "route_ab_balanced_results.jsonl"

LOW_CONF = 0.75
ACCEPT_CONF = 0.85
MID_CONF = 0.65
SMALL_MARGIN = 0.12
ROUTE_A_LLM_TOPK = 5
ROUTE_B_STRONG_CONF = 0.18

BUCKET_AGREE_HIGH = "AB一致高置信"
BUCKET_A_STRONG_B_WEAK = "A强B弱"
BUCKET_B_STRONG_A_WEAK = "B强A弱"
BUCKET_CONFLICT = "AB冲突"
BUCKET_BOTH_EMPTY = "双未命中"
BUCKET_LONG_OR_DAMAGED = "长文本/清洗异常"
BUCKET_LOW_REVIEW = "低置信待复核"

SAMPLED_BUCKET_TARGETS = {
    BUCKET_AGREE_HIGH: 20,
    BUCKET_A_STRONG_B_WEAK: 20,
    BUCKET_B_STRONG_A_WEAK: 20,
    BUCKET_CONFLICT: 50,
    BUCKET_BOTH_EMPTY: 50,
    BUCKET_LONG_OR_DAMAGED: 40,
}

ROUTE_B_LLM_PRIORITY = [
    BUCKET_CONFLICT,
    BUCKET_BOTH_EMPTY,
    BUCKET_LONG_OR_DAMAGED,
    BUCKET_B_STRONG_A_WEAK,
    BUCKET_A_STRONG_B_WEAK,
    BUCKET_LOW_REVIEW,
    BUCKET_AGREE_HIGH,
]

DETAIL_HEADERS = [
    "原sheet", "原序号", "拆分序号", "原始名称", "清洗后名称", "测试输入",
    "来源是否多产品行", "当前输入是否长文本", "当前输入是否清洗异常",
    "样本桶", "是否自动接受", "是否待人工复核",
    "是否进入LLM抽样", "LLM抽样原因", "LLM预算组", "LLM调用阶段",
    "Route B弱匹配标记",
    "A_local_node_id", "A_local_name", "A_local_path", "A_local_confidence",
    "A_local_source", "A_local_latency_ms", "A_top1_top2_margin",
    "B_local_node_id", "B_local_name", "B_local_path", "B_local_confidence",
    "B_local_source", "B_local_latency_ms", "B_local_layers",
    "A_LLM_node_id", "A_LLM_name", "A_LLM_path", "A_LLM_confidence", "A_LLM_source",
    "B_LLM_node_id", "B_LLM_name", "B_LLM_path", "B_LLM_confidence", "B_LLM_source",
    "A/B本地是否一致", "本地难例原因",
    "最终采用路线", "最终node_id", "最终name", "最终path", "最终confidence", "最终source",
    "复核建议",
]

EXTENSION_HEADERS = [
    "未命中产品名", "原sheet", "原序号", "原始名称", "清洗后名称",
    "最接近候选节点", "最接近候选路径", "候选分数",
    "建议动作", "建议新增节点名", "建议父节点node_id", "建议父节点路径",
    "建议同义词", "建议理由", "优先级", "复核状态",
]

EXTENSION_ACTIONS = {
    "清洗异常/不扩展",
    "补同义词",
    "新增叶子节点",
    "新增中间节点",
    "体系外/新增大类",
}

EXTENSION_SYSTEM = """你是产品标准体系维护助手。请判断一个未命中产品应该如何处理。
只能从这些建议动作中选择一个：清洗异常/不扩展、补同义词、新增叶子节点、新增中间节点、体系外/新增大类。
严格输出 JSON：
{"action":"<建议动作>","new_node_name":"<建议新增节点名或空>","parent_node_id":<int或null>,"parent_path":"<建议父节点路径或空>","synonyms":["<建议同义词>"],"reason":"<简短理由>","priority":"高/中/低"}"""


@dataclass
class Record:
    idx: int
    sheet: str
    seq: Any
    split_seq: int
    raw: str
    cleaned: str
    product: str
    is_split: bool


SPLIT_RE = re.compile(r"\s*[、；;]\s*")


def safe_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def split_products(text: str) -> list[str]:
    text = safe_text(text)
    if not text:
        return []
    parts = [p.strip(" \t\r\n,，。") for p in SPLIT_RE.split(text)]
    parts = [p for p in parts if p]
    return parts or [text]


def load_records(limit: int | None = None) -> list[Record]:
    wb = openpyxl.load_workbook(INPUT_XLSX, read_only=True, data_only=True)
    records: list[Record] = []
    try:
        idx = 0
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or all(v is None for v in row):
                    continue
                raw = safe_text(row[1] if len(row) > 1 else "")
                cleaned = safe_text(row[2] if len(row) > 2 else "")
                product = cleaned or raw
                if not product:
                    continue
                products = split_products(product)
                is_split = len(products) > 1
                for split_seq, item in enumerate(products, 1):
                    idx += 1
                    records.append(Record(
                        idx=idx,
                        sheet=sheet,
                        seq=row[0] if row and row[0] is not None else "",
                        split_seq=split_seq,
                        raw=raw,
                        cleaned=cleaned,
                        product=item,
                        is_split=is_split,
                    ))
                    if limit and len(records) >= limit:
                        return records
    finally:
        wb.close()
    return records


def empty_result(source: str = "") -> dict:
    return {
        "node_id": None,
        "name": None,
        "path": None,
        "confidence": 0.0,
        "source": source,
        "latency_ms": 0.0,
        "layers": 0,
    }


def as_result(result: dict | None) -> dict:
    result = result or {}
    return {
        "node_id": result.get("node_id"),
        "name": result.get("name"),
        "path": result.get("path"),
        "confidence": float(result.get("confidence", 0.0) or 0.0),
        "source": result.get("source", ""),
        "latency_ms": float(result.get("latency_ms", 0.0) or 0.0),
        "layers": int(result.get("n_layers_visited", result.get("layers", 0)) or 0),
    }


def route_a_local(mapper: ProductMapper, product: str) -> tuple[dict, float]:
    result, candidates = mapper.explain(product)
    margin = 0.0
    if len(candidates) >= 2:
        margin = float(candidates[0].get("fused", 0.0) - candidates[1].get("fused", 0.0))
    elif len(candidates) == 1:
        margin = float(candidates[0].get("fused", 0.0))
    return as_result(result), round(margin, 3)


def route_a_llm(mapper: ProductMapper, product: str) -> dict:
    return as_result(mapper.map(product, topk_candidates=ROUTE_A_LLM_TOPK))


def route_b(mapper: PageIndexMapper, product: str) -> dict:
    return as_result(mapper.map(product))


def has_multiple_products(text: str) -> bool:
    if len(text) >= 80 and re.search(r"[、；;]", text):
        return True
    return len(re.findall(r"[、；;]", text)) >= 2


def is_long_input(text: str) -> bool:
    return len(text) >= 80 or has_multiple_products(text)


def looks_damaged(product: str) -> bool:
    product = product.strip()
    if len(product) <= 2:
        return True
    if re.fullmatch(r"[-A-Za-z0-9._/+\\ ]{1,20}", product):
        return True
    if "?" in product or "锟" in product:
        return True
    return False


def hit(result: dict) -> bool:
    return result.get("node_id") is not None


def route_a_strong(a: dict, margin: float) -> bool:
    return hit(a) and (
        a["confidence"] >= ACCEPT_CONF or
        (a["confidence"] >= MID_CONF and margin >= SMALL_MARGIN)
    )


def route_b_strong(b: dict) -> bool:
    return hit(b) and (
        b["source"] == "pageindex_exact" or
        (b["source"] == "pageindex_trigram" and b["confidence"] >= ROUTE_B_STRONG_CONF) or
        (b["source"] == "pageindex" and b["confidence"] >= LOW_CONF)
    )


def route_b_weak(b: dict) -> bool:
    return b.get("source") == "pageindex_trigram_weak"


def local_agree(a: dict, b: dict) -> bool:
    return hit(a) and hit(b) and a["node_id"] == b["node_id"]


def local_conflict(a: dict, b: dict) -> bool:
    return hit(a) and hit(b) and a["node_id"] != b["node_id"]


def local_reasons(record: Record, a: dict, b: dict, margin: float) -> list[str]:
    reasons: list[str] = []
    if not hit(a):
        reasons.append("Route A未命中")
    if not hit(b):
        reasons.append("Route B未命中")
    if local_conflict(a, b):
        reasons.append("A/B节点不一致")
    if hit(a) and a["confidence"] < MID_CONF:
        reasons.append("Route A低置信")
    if hit(b) and b["confidence"] < LOW_CONF:
        reasons.append("Route B低置信")
    if hit(a) and a["confidence"] < ACCEPT_CONF and margin < SMALL_MARGIN:
        reasons.append("Route A候选分差小")
    if is_long_input(record.product):
        reasons.append("当前输入长文本")
    if looks_damaged(record.product):
        reasons.append("产品名疑似清洗损坏")
    if route_b_weak(b):
        reasons.append("Route B弱匹配")
    return reasons


def is_auto_accept(a: dict, b: dict, margin: float) -> bool:
    return (
        local_agree(a, b) or
        (hit(a) and a["confidence"] >= ACCEPT_CONF) or
        b["source"] == "pageindex_exact" or
        (b["source"] == "pageindex_trigram" and b["confidence"] >= ROUTE_B_STRONG_CONF) or
        (hit(a) and a["confidence"] >= MID_CONF and margin >= SMALL_MARGIN)
    )


def classify_bucket(record: Record, a: dict, b: dict, margin: float) -> str:
    if is_long_input(record.product) or looks_damaged(record.product):
        return BUCKET_LONG_OR_DAMAGED
    if not hit(a) and not hit(b):
        return BUCKET_BOTH_EMPTY
    if local_conflict(a, b):
        return BUCKET_CONFLICT
    if local_agree(a, b):
        return BUCKET_AGREE_HIGH
    a_strong = route_a_strong(a, margin)
    b_strong = route_b_strong(b)
    if a_strong and not b_strong:
        return BUCKET_A_STRONG_B_WEAK
    if b_strong and not a_strong:
        return BUCKET_B_STRONG_A_WEAK
    return BUCKET_LOW_REVIEW


def choose_final(row: dict) -> tuple[str, dict]:
    a = row["A_local"]
    b = row["B_local"]
    a_llm = row["A_LLM"]
    b_llm = row["B_LLM"]
    margin = row["A_top1_top2_margin"]

    if hit(a_llm) and a_llm["confidence"] >= LOW_CONF:
        if hit(b_llm) and b_llm["confidence"] > a_llm["confidence"]:
            return "Route B LLM", b_llm
        return "Route A LLM", a_llm
    if hit(b_llm) and b_llm["confidence"] >= LOW_CONF:
        return "Route B LLM", b_llm
    if local_agree(a, b):
        return "A/B一致本地", a
    if route_a_strong(a, margin):
        return "Route A Local", a
    if route_b_strong(b):
        return "Route B Local", b
    if hit(a) and (not hit(b) or a["confidence"] >= b["confidence"]):
        return "Route A Local", a
    if hit(b):
        return "Route B Local", b
    return "Empty", empty_result("empty")


def review_suggestion(row: dict) -> str:
    final = row["最终"]
    a = row["A_local"]
    b = row["B_local"]
    if not hit(final):
        return "未命中，建议人工复核或考虑体系扩展"
    if row["是否待人工复核"] == "否":
        return "可接受"
    if local_conflict(a, b):
        return "A/B本地冲突，建议人工复核"
    if route_b_weak(b):
        return "Route B弱匹配，建议人工复核"
    if row["样本桶"] == BUCKET_LONG_OR_DAMAGED:
        return "长文本或清洗异常，建议人工复核"
    if final["confidence"] < LOW_CONF:
        return "最终结果置信度偏低，建议人工复核"
    return "进入低置信/抽样检查集合，建议抽查"


def build_local_rows(records: list[Record], mapper_a: ProductMapper, mapper_b: PageIndexMapper,
                     original_key: str) -> list[dict]:
    rows: list[dict] = []
    config.DEEPSEEK_API_KEY = ""
    for i, record in enumerate(records, 1):
        a_local, margin = route_a_local(mapper_a, record.product)
        b_local = route_b(mapper_b, record.product)
        bucket = classify_bucket(record, a_local, b_local, margin)
        reasons = local_reasons(record, a_local, b_local, margin)
        auto_accept = is_auto_accept(a_local, b_local, margin)
        row = {
            "原sheet": record.sheet,
            "原序号": record.seq,
            "拆分序号": record.split_seq,
            "原始名称": record.raw,
            "清洗后名称": record.cleaned,
            "测试输入": record.product,
            "来源是否多产品行": "是" if record.is_split else "否",
            "当前输入是否长文本": "是" if is_long_input(record.product) else "否",
            "当前输入是否清洗异常": "是" if looks_damaged(record.product) else "否",
            "样本桶": bucket,
            "是否自动接受": "是" if auto_accept else "否",
            "是否待人工复核": "否" if auto_accept else "是",
            "是否进入LLM抽样": "否",
            "LLM抽样原因": "",
            "LLM预算组": "",
            "LLM调用阶段": "",
            "Route B弱匹配标记": "是" if route_b_weak(b_local) else "否",
            "A_local": a_local,
            "B_local": b_local,
            "A_LLM": empty_result(),
            "B_LLM": empty_result(),
            "A/B本地是否一致": "是" if local_agree(a_local, b_local) else "否",
            "本地难例原因": "；".join(reasons),
            "A_top1_top2_margin": margin,
        }
        final_route, final = choose_final(row)
        row["最终采用路线"] = final_route
        row["最终"] = final
        row["复核建议"] = review_suggestion(row)
        rows.append(row)

        if i % 100 == 0:
            print(f"  本地阶段已处理 {i}/{len(records)}")

    config.DEEPSEEK_API_KEY = original_key
    return rows


def stratified_sample_rows(rows: list[dict], sample_size: int, sample_seed: int) -> list[dict]:
    if sample_size <= 0 or sample_size >= len(rows):
        return rows

    rng = random.Random(sample_seed)
    by_bucket: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        by_bucket[row["样本桶"]].append(row)

    bucket_order = [
        BUCKET_AGREE_HIGH,
        BUCKET_A_STRONG_B_WEAK,
        BUCKET_B_STRONG_A_WEAK,
        BUCKET_CONFLICT,
        BUCKET_BOTH_EMPTY,
        BUCKET_LONG_OR_DAMAGED,
        BUCKET_LOW_REVIEW,
    ]
    active_buckets = [b for b in bucket_order if by_bucket.get(b)]
    if not active_buckets:
        return rng.sample(rows, sample_size)

    selected: list[dict] = []
    selected_ids: set[int] = set()
    base_take = max(1, sample_size // len(active_buckets))
    for bucket in active_buckets:
        candidates = by_bucket[bucket]
        take = min(base_take, len(candidates), sample_size - len(selected))
        if take <= 0:
            continue
        picked = rng.sample(candidates, take)
        selected.extend(picked)
        selected_ids.update(id(row) for row in picked)

    if len(selected) < sample_size:
        remaining = [row for row in rows if id(row) not in selected_ids]
        selected.extend(rng.sample(remaining, min(sample_size - len(selected), len(remaining))))

    selected.sort(key=lambda r: (str(r["原sheet"]), str(r["原序号"]), int(r["拆分序号"] or 0)))
    return selected


def apply_sampling(rows: list[dict], sample_mode: str, sample_size: int | None,
                   sample_seed: int) -> list[dict]:
    if sample_size is None:
        return rows
    if sample_mode == "first":
        return rows[:sample_size]
    if sample_mode == "stratified":
        return stratified_sample_rows(rows, sample_size, sample_seed)
    raise ValueError(f"unknown sample mode: {sample_mode}")


def scaled_bucket_targets(llm_budget: int) -> dict[str, int]:
    if llm_budget <= 0:
        return {}
    base_total = sum(SAMPLED_BUCKET_TARGETS.values())
    targets: dict[str, int] = {}
    used = 0
    for bucket, target in SAMPLED_BUCKET_TARGETS.items():
        value = int(target * llm_budget / base_total)
        targets[bucket] = value
        used += value
    priority = [
        BUCKET_CONFLICT,
        BUCKET_BOTH_EMPTY,
        BUCKET_LONG_OR_DAMAGED,
        BUCKET_A_STRONG_B_WEAK,
        BUCKET_B_STRONG_A_WEAK,
        BUCKET_AGREE_HIGH,
    ]
    for bucket in priority:
        if used >= llm_budget:
            break
        targets[bucket] = targets.get(bucket, 0) + 1
        used += 1
    return targets


def pick_sampled_indexes(rows: list[dict], llm_budget: int, sample_seed: int) -> set[int]:
    rng = random.Random(sample_seed)
    by_bucket: dict[str, list[int]] = defaultdict(list)
    for idx, row in enumerate(rows):
        by_bucket[row["样本桶"]].append(idx)

    selected: set[int] = set()
    targets = scaled_bucket_targets(llm_budget)
    for bucket, target in targets.items():
        candidates = by_bucket.get(bucket, [])
        if not candidates or target <= 0:
            continue
        selected.update(rng.sample(candidates, min(target, len(candidates))))

    if len(selected) < llm_budget:
        fallback = [idx for idx in by_bucket.get(BUCKET_LOW_REVIEW, []) if idx not in selected]
        take = min(llm_budget - len(selected), len(fallback))
        if take > 0:
            selected.update(rng.sample(fallback, take))

    if len(selected) < llm_budget:
        fallback = [idx for idx in range(len(rows)) if idx not in selected]
        take = min(llm_budget - len(selected), len(fallback))
        if take > 0:
            selected.update(rng.sample(fallback, take))

    return selected


def pick_route_b_llm_indexes(rows: list[dict], selected: set[int], route_b_llm_budget: int) -> set[int]:
    if route_b_llm_budget <= 0:
        return set()
    picked: list[int] = []
    for bucket in ROUTE_B_LLM_PRIORITY:
        for idx in sorted(selected):
            if len(picked) >= route_b_llm_budget:
                return set(picked)
            if rows[idx]["样本桶"] == bucket and idx not in picked:
                picked.append(idx)
    return set(picked)


def indexes_for_strategy(rows: list[dict], strategy: str, llm_budget: int,
                         sample_seed: int) -> tuple[set[int], set[int]]:
    if strategy == "none":
        return set(), set()
    if strategy == "sampled":
        selected = pick_sampled_indexes(rows, llm_budget, sample_seed)
        return selected, set()
    all_rows = set(range(len(rows)))
    if strategy == "full-a":
        return all_rows, set()
    if strategy == "full-b":
        return set(), all_rows
    if strategy == "full-ab":
        return all_rows, all_rows
    raise ValueError(f"unknown llm strategy: {strategy}")


def apply_llm_stage(rows: list[dict], mapper_a: ProductMapper, mapper_b: PageIndexMapper,
                    original_key: str, strategy: str, llm_budget: int,
                    route_b_llm: bool, route_b_llm_budget: int,
                    sample_seed: int) -> None:
    a_llm_indexes, b_llm_indexes = indexes_for_strategy(rows, strategy, llm_budget, sample_seed)
    if strategy == "sampled" and route_b_llm:
        b_llm_indexes = pick_route_b_llm_indexes(rows, a_llm_indexes, route_b_llm_budget)

    total_llm_calls = len(a_llm_indexes) + len(b_llm_indexes)
    if total_llm_calls == 0:
        return

    config.DEEPSEEK_API_KEY = original_key
    done = 0
    for idx, row in enumerate(rows):
        stages: list[str] = []
        if idx in a_llm_indexes:
            done += 1
            row["是否进入LLM抽样"] = "是"
            row["LLM抽样原因"] = row["样本桶"]
            row["LLM预算组"] = row["样本桶"] if strategy == "sampled" else strategy
            print(f"  LLM {done}/{total_llm_calls} | Route A | {row['测试输入'][:40]}", flush=True)
            row["A_LLM"] = route_a_llm(mapper_a, row["测试输入"])
            stages.append("Route A LLM")

        if idx in b_llm_indexes:
            done += 1
            row["是否进入LLM抽样"] = "是"
            row["LLM抽样原因"] = row["样本桶"]
            row["LLM预算组"] = row["样本桶"] if strategy == "sampled" else strategy
            print(f"  LLM {done}/{total_llm_calls} | Route B | {row['测试输入'][:40]}", flush=True)
            row["B_LLM"] = route_b(mapper_b, row["测试输入"])
            stages.append("Route B LLM")

        if stages:
            row["LLM调用阶段"] = " + ".join(stages)
            final_route, final = choose_final(row)
            row["最终采用路线"] = final_route
            row["最终"] = final
            if hit(final) and final["confidence"] >= LOW_CONF and not local_conflict(row["A_local"], row["B_local"]):
                row["是否待人工复核"] = "否"
            row["复核建议"] = review_suggestion(row)

    config.DEEPSEEK_API_KEY = ""


def extension_triggered(row: dict) -> bool:
    called_a = "Route A LLM" in row.get("LLM调用阶段", "")
    called_b = "Route B LLM" in row.get("LLM调用阶段", "")
    if called_a or called_b:
        return (not hit(row["A_LLM"])) and (not hit(row["B_LLM"]))
    return (not hit(row["A_local"])) and (not hit(row["B_local"]))


def nearest_taxonomy_candidates(mapper_a: ProductMapper, product: str, top_k: int = 5) -> list[dict]:
    original_key = config.DEEPSEEK_API_KEY
    config.DEEPSEEK_API_KEY = ""
    try:
        cands = mapper_a.recaller.recall(product)
        ordered = _fuse(cands)[:top_k]
    finally:
        config.DEEPSEEK_API_KEY = original_key

    items = []
    for cand in ordered:
        node = cand.node
        parent_id = node.parent_id if node.is_leaf else node.id
        if node.is_leaf and len(node.path_names) > 1:
            parent_path = " > ".join(node.path_names[:-1])
        else:
            parent_path = node.path_str
        items.append({
            "node_id": node.id,
            "name": node.name,
            "path": node.path_str,
            "parent_id": parent_id,
            "parent_path": parent_path,
            "is_leaf": node.is_leaf,
            "score": round(float(getattr(cand, "fused", 0.0)), 3),
            "trgm": round(float(getattr(cand, "trgm", 0.0)), 3),
            "vec": round(float(getattr(cand, "vec", 0.0)), 3),
        })
    return items


def format_extension_candidates(candidates: list[dict]) -> str:
    lines = []
    for i, item in enumerate(candidates, 1):
        lines.append(
            f"{i}. node_id={item['node_id']} | {item['path']} | "
            f"score={item['score']}, trgm={item['trgm']}, vec={item['vec']}"
        )
    return "\n".join(lines) if lines else "无候选"


def normalize_extension_output(out: dict | None, product: str, candidates: list[dict]) -> dict:
    best = candidates[0] if candidates else {}
    if not out:
        if looks_damaged(product):
            action = "清洗异常/不扩展"
            priority = "低"
            reason = "产品名过短、疑似型号或清洗后存在异常字符，先复核清洗结果"
        elif best and best.get("score", 0.0) >= 1.0:
            action = "补同义词"
            priority = "中"
            reason = "存在相近标准节点，优先作为同义词候选人工确认"
        elif best and best.get("score", 0.0) >= 0.45:
            action = "新增叶子节点"
            priority = "中"
            reason = "存在相近父路径，但当前体系缺少更细产品节点"
        else:
            action = "体系外/新增大类"
            priority = "低"
            reason = "没有足够相近的候选路径，建议人工判断是否属于体系外领域"
        return {
            "action": action,
            "new_node_name": "" if action in {"清洗异常/不扩展", "补同义词"} else product,
            "parent_node_id": best.get("parent_id"),
            "parent_path": best.get("parent_path", ""),
            "synonyms": [product] if action == "补同义词" else [],
            "reason": reason,
            "priority": priority,
        }

    action = str(out.get("action") or "").strip()
    if action not in EXTENSION_ACTIONS:
        action = "体系外/新增大类"
    synonyms = out.get("synonyms") or []
    if not isinstance(synonyms, list):
        synonyms = [str(synonyms)]
    parent_node_id = out.get("parent_node_id")
    try:
        parent_node_id = int(parent_node_id) if parent_node_id not in (None, "") else None
    except Exception:
        parent_node_id = None
    priority = str(out.get("priority") or "中").strip()
    if priority not in {"高", "中", "低"}:
        priority = "中"
    return {
        "action": action,
        "new_node_name": str(out.get("new_node_name") or "").strip(),
        "parent_node_id": parent_node_id,
        "parent_path": str(out.get("parent_path") or "").strip(),
        "synonyms": [str(x).strip() for x in synonyms if str(x).strip()],
        "reason": str(out.get("reason") or "").strip(),
        "priority": priority,
    }


def extension_llm_suggest(product: str, raw: str, cleaned: str,
                          candidates: list[dict], original_key: str) -> dict | None:
    if not original_key:
        return None
    user = (
        f"未命中产品名：{product}\n"
        f"原始名称：{raw}\n"
        f"清洗后名称：{cleaned}\n\n"
        f"最接近候选节点 Top5：\n{format_extension_candidates(candidates)}"
    )
    config.DEEPSEEK_API_KEY = original_key
    try:
        return chat_json(EXTENSION_SYSTEM, user, timeout=60)
    finally:
        config.DEEPSEEK_API_KEY = ""


def build_extension_suggestions(rows: list[dict], mapper_a: ProductMapper,
                                original_key: str) -> list[dict]:
    suggestions: list[dict] = []
    for row in rows:
        if not extension_triggered(row):
            continue
        product = row["测试输入"]
        candidates = nearest_taxonomy_candidates(mapper_a, product, top_k=5)
        llm_out = extension_llm_suggest(product, row["原始名称"], row["清洗后名称"],
                                        candidates, original_key)
        suggestion = normalize_extension_output(llm_out, product, candidates)
        best = candidates[0] if candidates else {}
        suggestions.append({
            "未命中产品名": product,
            "原sheet": row["原sheet"],
            "原序号": row["原序号"],
            "原始名称": row["原始名称"],
            "清洗后名称": row["清洗后名称"],
            "最接近候选节点": best.get("name", ""),
            "最接近候选路径": best.get("path", ""),
            "候选分数": best.get("score", ""),
            "建议动作": suggestion["action"],
            "建议新增节点名": suggestion["new_node_name"],
            "建议父节点node_id": suggestion["parent_node_id"] if suggestion["parent_node_id"] is not None else "",
            "建议父节点路径": suggestion["parent_path"],
            "建议同义词": "、".join(suggestion["synonyms"]),
            "建议理由": suggestion["reason"],
            "优先级": suggestion["priority"],
            "复核状态": "待复核",
        })
    config.DEEPSEEK_API_KEY = original_key
    return suggestions


def flatten_extension(row: dict) -> list:
    return [row.get(header, "") for header in EXTENSION_HEADERS]


def estimate_llm_call_count(rows: list[dict], extension_rows: list[dict]) -> int:
    count = 0
    for row in rows:
        stage = row.get("LLM调用阶段", "")
        if "Route A LLM" in stage and row["A_LLM"].get("source") != "exact_match":
            count += 1
        if "Route B LLM" in stage and row["B_LLM"].get("source") != "pageindex_exact":
            count += 1
    count += len(extension_rows)
    return count


def estimate_token_cost(rows: list[dict], extension_rows: list[dict]) -> int:
    # Rough estimate for planning/reporting only. DeepSeek usage is not exposed by llm.py yet.
    route_a_calls = 0
    route_b_calls = 0
    for row in rows:
        stage = row.get("LLM调用阶段", "")
        if "Route A LLM" in stage and row["A_LLM"].get("source") != "exact_match":
            route_a_calls += 1
        if "Route B LLM" in stage and row["B_LLM"].get("source") != "pageindex_exact":
            route_b_calls += 1
    return route_a_calls * 1800 + route_b_calls * 4500 + len(extension_rows) * 1200


def flatten_detail(row: dict) -> list:
    a, b = row["A_local"], row["B_local"]
    a_llm, b_llm = row["A_LLM"], row["B_LLM"]
    final = row["最终"]
    return [
        row["原sheet"], row["原序号"], row["拆分序号"], row["原始名称"], row["清洗后名称"], row["测试输入"],
        row["来源是否多产品行"], row["当前输入是否长文本"], row["当前输入是否清洗异常"],
        row["样本桶"], row["是否自动接受"], row["是否待人工复核"],
        row["是否进入LLM抽样"], row["LLM抽样原因"], row["LLM预算组"], row["LLM调用阶段"],
        row["Route B弱匹配标记"],
        a["node_id"], a["name"], a["path"], a["confidence"], a["source"], a["latency_ms"],
        row["A_top1_top2_margin"],
        b["node_id"], b["name"], b["path"], b["confidence"], b["source"], b["latency_ms"], b["layers"],
        a_llm["node_id"], a_llm["name"], a_llm["path"], a_llm["confidence"], a_llm["source"],
        b_llm["node_id"], b_llm["name"], b_llm["path"], b_llm["confidence"], b_llm["source"],
        row["A/B本地是否一致"], row["本地难例原因"],
        row["最终采用路线"], final["node_id"], final["name"], final["path"], final["confidence"], final["source"],
        row["复核建议"],
    ]


def append_jsonl(rows: list[dict], jsonl_path: Path) -> None:
    jsonl_path.parent.mkdir(exist_ok=True)
    with jsonl_path.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False, default=str) + "\n")


def append_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def autosize(ws, max_width: int = 60) -> None:
    for col_idx, cells in enumerate(ws.columns, 1):
        width = 10
        for cell in cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(max_width, len(value) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def ratio(numerator: int | float, denominator: int | float) -> float:
    return float(numerator) / float(denominator) if denominator else 0.0


def summarize(rows: list[dict], extension_rows: list[dict], llm_stats: dict,
              started: float, embedder: str, strategy: str, llm_budget: int, route_b_llm: bool,
              route_b_llm_budget: int, sample_seed: int) -> list[tuple]:
    n = len(rows)
    source_rows = len({(r["原sheet"], r["原序号"]) for r in rows})
    a_hits = sum(1 for r in rows if hit(r["A_local"]))
    b_hits = sum(1 for r in rows if hit(r["B_local"]))
    a_llm_hit = sum(1 for r in rows if hit(r["A_LLM"]))
    b_llm_hit = sum(1 for r in rows if hit(r["B_LLM"]))
    final_hits = sum(1 for r in rows if hit(r["最终"]))
    both_hit = [r for r in rows if hit(r["A_local"]) and hit(r["B_local"])]
    agree = sum(1 for r in both_hit if r["A_local"]["node_id"] == r["B_local"]["node_id"])
    conflict = sum(1 for r in both_hit if r["A_local"]["node_id"] != r["B_local"]["node_id"])
    both_llm_hit = [r for r in rows if hit(r["A_LLM"]) and hit(r["B_LLM"])]
    llm_agree = sum(1 for r in both_llm_hit if r["A_LLM"]["node_id"] == r["B_LLM"]["node_id"])
    llm_conflict = sum(1 for r in both_llm_hit if r["A_LLM"]["node_id"] != r["B_LLM"]["node_id"])
    double_miss = sum(1 for r in rows if extension_triggered(r))
    llm_rows = sum(1 for r in rows if r["是否进入LLM抽样"] == "是")
    llm_call_count = int(llm_stats.get("requests", 0) or 0)
    estimated_tokens = estimate_token_cost(rows, extension_rows)
    auto_accept = sum(1 for r in rows if r["是否自动接受"] == "是")
    review = sum(1 for r in rows if r["是否待人工复核"] == "是")
    b_source_counts = Counter(r["B_local"]["source"] for r in rows)
    return [
        ("原始行数", source_rows, "清洗表原始记录数"),
        ("测试产品项数", n, "多产品行拆分后的测试样本数"),
        ("自动接受数量", auto_accept, ""),
        ("自动接受比例", ratio(auto_accept, n), ""),
        ("待人工复核数量", review, ""),
        ("待人工复核比例", ratio(review, n), ""),
        ("LLM调用样本数", llm_rows, "至少调用过 Route A 或 Route B LLM 的样本数"),
        ("LLM调用比例", ratio(llm_rows, n), ""),
        ("Route A本地命中率", ratio(a_hits, n), ""),
        ("Route B本地命中率", ratio(b_hits, n), ""),
        ("Route A完整命中率", ratio(a_llm_hit, n), "统计 A_LLM 结果；未启用 LLM 时为 0"),
        ("Route B完整命中率", ratio(b_llm_hit, n), "统计 B_LLM 结果；未启用 LLM 时为 0"),
        ("Hybrid最终命中率", ratio(final_hits, n), ""),
        ("A/B本地一致率", ratio(agree, len(both_hit)), "仅统计 A/B 都命中的样本"),
        ("A/B本地冲突率", ratio(conflict, len(both_hit)), "仅统计 A/B 都命中的样本"),
        ("A/B完整一致率", ratio(llm_agree, len(both_llm_hit)), "仅统计 A_LLM/B_LLM 都命中的样本"),
        ("A/B完整冲突率", ratio(llm_conflict, len(both_llm_hit)), "仅统计 A_LLM/B_LLM 都命中的样本"),
        ("双路线未命中数量", double_miss, "完整 LLM 模式优先看 A_LLM+B_LLM，非 LLM 模式看本地 A+B"),
        ("体系扩展建议数量", len(extension_rows), ""),
        ("Route B精确匹配比例", ratio(b_source_counts.get("pageindex_exact", 0), n), ""),
        ("Route B强候选比例", ratio(b_source_counts.get("pageindex_trigram", 0), n), ""),
        ("Route B弱候选比例", ratio(b_source_counts.get("pageindex_trigram_weak", 0), n), ""),
        ("Route B空结果比例", ratio(b_source_counts.get("pageindex_empty", 0), n), ""),
        ("Route A平均本地置信度", sum(r["A_local"]["confidence"] for r in rows) / max(1, n), ""),
        ("Route B平均本地置信度", sum(r["B_local"]["confidence"] for r in rows) / max(1, n), ""),
        ("Route A平均本地延迟ms", sum(r["A_local"]["latency_ms"] for r in rows) / max(1, n), ""),
        ("Route B平均本地延迟ms", sum(r["B_local"]["latency_ms"] for r in rows) / max(1, n), ""),
        ("Route A平均完整延迟ms", sum(r["A_LLM"]["latency_ms"] for r in rows if r["A_LLM"]["source"]) / max(1, sum(1 for r in rows if r["A_LLM"]["source"])), ""),
        ("Route B平均完整延迟ms", sum(r["B_LLM"]["latency_ms"] for r in rows if r["B_LLM"]["source"]) / max(1, sum(1 for r in rows if r["B_LLM"]["source"])), ""),
        ("LLM实际请求次数", llm_call_count, "由底层 LLM 客户端统计；Route B 树搜索可能一次样本多次请求"),
        ("LLM失败次数", int(llm_stats.get("failures", 0) or 0), ""),
        ("LLM返回真实tokens", int(llm_stats.get("total_tokens", 0) or 0), "API 未返回 usage 时为 0"),
        ("LLM请求字符数", int(llm_stats.get("prompt_chars", 0) or 0), ""),
        ("token成本估算", estimated_tokens, "粗估：Route A 1800/次，Route B 4500/次，体系扩展 1200/次"),
        ("总耗时秒", time.time() - started, ""),
        ("Route A embedder", embedder, ""),
        ("LLM strategy", strategy, ""),
        ("LLM budget", llm_budget, ""),
        ("Route B LLM树搜索", "启用" if route_b_llm else "禁用", ""),
        ("Route B LLM budget", route_b_llm_budget, ""),
        ("sample seed", sample_seed, ""),
    ]


def write_bucket_stats(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("样本桶统计")
    append_header(ws, ["样本桶", "数量", "占比", "LLM抽样数量", "自动接受数量", "待复核数量", "A平均置信度", "B平均置信度"])
    by_bucket: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        by_bucket[row["样本桶"]].append(row)
    total = len(rows)
    for bucket, items in sorted(by_bucket.items(), key=lambda x: (-len(x[1]), x[0])):
        n = len(items)
        ws.append([
            bucket,
            n,
            ratio(n, total),
            sum(1 for r in items if r["是否进入LLM抽样"] == "是"),
            sum(1 for r in items if r["是否自动接受"] == "是"),
            sum(1 for r in items if r["是否待人工复核"] == "是"),
            sum(r["A_local"]["confidence"] for r in items) / max(1, n),
            sum(r["B_local"]["confidence"] for r in items) / max(1, n),
        ])


def write_excel(rows: list[dict], extension_rows: list[dict], llm_stats: dict,
                output_path: Path, jsonl_path: Path, started: float, embedder: str, strategy: str,
                llm_budget: int, route_b_llm: bool, route_b_llm_budget: int,
                sample_seed: int, sample_mode: str, sample_size: int | None) -> None:
    output_path.parent.mkdir(exist_ok=True)
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "实验汇总"
    append_header(ws_summary, ["指标", "数值", "说明"])
    for item in summarize(rows, extension_rows, llm_stats, started, embedder, strategy,
                          llm_budget, route_b_llm, route_b_llm_budget, sample_seed):
        ws_summary.append(item)

    ws_detail = wb.create_sheet("全量本地明细")
    append_header(ws_detail, DETAIL_HEADERS)
    for row in rows:
        ws_detail.append(flatten_detail(row))

    write_bucket_stats(wb, rows)

    ws_llm = wb.create_sheet("LLM抽样样本")
    append_header(ws_llm, DETAIL_HEADERS)
    for row in rows:
        if row["是否进入LLM抽样"] == "是":
            ws_llm.append(flatten_detail(row))

    ws_route_b_llm = wb.create_sheet("RouteB树搜索样本")
    append_header(ws_route_b_llm, DETAIL_HEADERS)
    for row in rows:
        if row["B_LLM"]["source"] == "pageindex":
            ws_route_b_llm.append(flatten_detail(row))

    ws_conflict = wb.create_sheet("冲突样本")
    append_header(ws_conflict, DETAIL_HEADERS)
    for row in rows:
        if local_conflict(row["A_local"], row["B_local"]):
            ws_conflict.append(flatten_detail(row))

    ws_review = wb.create_sheet("待复核样本")
    append_header(ws_review, DETAIL_HEADERS)
    for row in rows:
        if row["是否待人工复核"] == "是":
            ws_review.append(flatten_detail(row))

    ws_extension = wb.create_sheet("体系扩展建议")
    append_header(ws_extension, EXTENSION_HEADERS)
    for row in extension_rows:
        ws_extension.append(flatten_extension(row))

    ws_dist = wb.create_sheet("来源分布")
    append_header(ws_dist, ["路线", "source", "数量", "占比"])
    for route, key in [("Route A Local", "A_local"), ("Route B Local", "B_local"),
                       ("Route A LLM", "A_LLM"), ("Route B LLM", "B_LLM"),
                       ("最终", "最终")]:
        counts = Counter(row[key]["source"] for row in rows if row[key]["source"])
        total = sum(counts.values()) or 1
        for source, count in counts.most_common():
            ws_dist.append([route, source, count, count / total])

    ws_config = wb.create_sheet("实验配置")
    append_header(ws_config, ["配置项", "值"])
    config_rows = [
        ("输入文件", str(INPUT_XLSX)),
        ("输出文件", str(output_path)),
        ("中间JSONL", str(jsonl_path)),
        ("Route A embedder", embedder),
        ("LOW_CONF", LOW_CONF),
        ("ACCEPT_CONF", ACCEPT_CONF),
        ("MID_CONF", MID_CONF),
        ("SMALL_MARGIN", SMALL_MARGIN),
        ("ROUTE_B_STRONG_CONF", ROUTE_B_STRONG_CONF),
        ("Route A LLM TopK", ROUTE_A_LLM_TOPK),
        ("LLM strategy", strategy),
        ("LLM budget", llm_budget),
        ("sample mode", sample_mode),
        ("sample size", "" if sample_size is None else sample_size),
        ("Route B LLM树搜索", "启用" if route_b_llm else "禁用"),
        ("Route B LLM budget", route_b_llm_budget),
        ("sample seed", sample_seed),
        ("LLM actual requests", int(llm_stats.get("requests", 0) or 0)),
        ("LLM failures", int(llm_stats.get("failures", 0) or 0)),
        ("LLM total_tokens", int(llm_stats.get("total_tokens", 0) or 0)),
        ("K_TRGM", config.K_TRGM),
        ("K_VEC", config.K_VEC),
        ("K_RERANK", config.K_RERANK),
        ("DeepSeek model", config.DEEPSEEK_MODEL),
        ("说明", "不计算准确率，只评估覆盖、一致、冲突、置信度、复核量和LLM调用比例"),
    ]
    for item in config_rows:
        ws_config.append(item)

    for ws in wb.worksheets:
        autosize(ws)
    wb.save(output_path)


def normalize_strategy(args: argparse.Namespace, original_key: str) -> str:
    if args.no_llm:
        return "none"
    strategy = args.llm_strategy
    if not original_key:
        return "none"
    return strategy


def resolve_output_path(raw: str | None) -> Path:
    if not raw:
        return OUTPUT_XLSX
    path = Path(raw)
    if not path.is_absolute():
        path = PROJECT_ROOT / path
    return path


def main() -> None:
    parser = argparse.ArgumentParser(description="Low-token balanced Route A / Route B evaluation")
    parser.add_argument("--limit", type=int, default=None, help="limit split product items for smoke tests")
    parser.add_argument("--sample-mode", choices=["none", "first", "stratified"], default="none",
                        help="optional sampling after local A/B bucketing")
    parser.add_argument("--sample-size", type=int, default=None,
                        help="number of split product items kept after sampling")
    parser.add_argument("--output", type=str, default=None, help="output Excel path")
    parser.add_argument("--llm-strategy", choices=["none", "sampled", "full-a", "full-b", "full-ab"],
                        default="sampled", help="LLM usage strategy")
    parser.add_argument("--llm-budget", type=int, default=200, help="sampled strategy LLM sample budget")
    parser.add_argument("--route-b-llm-budget", type=int, default=80,
                        help="maximum Route B LLM tree-search samples in sampled mode")
    parser.add_argument("--sample-seed", type=int, default=42, help="deterministic sampled LLM seed")
    parser.add_argument("--no-llm", action="store_true", help="deprecated alias for --llm-strategy none")
    parser.add_argument("--max-llm", type=int, default=None,
                        help="deprecated alias for --llm-budget in sampled mode")
    parser.add_argument("--route-b-llm", action="store_true",
                        help="also show sampled Route B PageIndex LLM tree search")
    parser.add_argument("--embedder", choices=["st", "hash"], default="st", help="Route A embedder")
    args = parser.parse_args()

    if args.max_llm is not None and args.llm_strategy == "sampled":
        args.llm_budget = args.max_llm

    if args.embedder == "st" and not st_available():
        print("[提示] sentence-transformers 未安装，自动改用 hash embedder。")
        args.embedder = "hash"

    original_key = config.DEEPSEEK_API_KEY
    strategy = normalize_strategy(args, original_key)
    if strategy == "none" and args.llm_strategy != "none":
        print("[提示] 未检测到可用 API key 或已禁用 LLM，本次只跑全量本地 A/B。")

    output_path = resolve_output_path(args.output)
    jsonl_path = output_path.with_suffix(".jsonl")
    if jsonl_path.exists():
        jsonl_path.unlink()

    started = time.time()
    print("读取清洗数据...")
    records = load_records(limit=args.limit)
    print(f"测试产品项数: {len(records)}")

    print("加载标准体系与映射器...")
    config.RECALL_BACKEND = "memory"
    config.EMBEDDER = args.embedder
    nodes = load_nodes()
    mapper_a = ProductMapper(nodes)
    if args.embedder != mapper_a.embedder_type:
        mapper_a.set_embedder(args.embedder)
    mapper_b = PageIndexMapper(nodes)

    print(f"第一阶段: 全量本地 A/B，embedder={args.embedder}")
    rows = build_local_rows(records, mapper_a, mapper_b, original_key)
    if args.sample_mode != "none" and args.sample_size:
        before = len(rows)
        rows = apply_sampling(rows, args.sample_mode, args.sample_size, args.sample_seed)
        print(f"抽样完成: {before} -> {len(rows)}，mode={args.sample_mode}, seed={args.sample_seed}")

    reset_llm_stats()
    print(
        f"第二阶段: LLM strategy={strategy}, llm_budget={args.llm_budget}, "
        f"route_b_llm={'启用' if args.route_b_llm else '禁用'}"
    )
    apply_llm_stage(rows, mapper_a, mapper_b, original_key, strategy, args.llm_budget,
                    args.route_b_llm, args.route_b_llm_budget, args.sample_seed)

    print("生成体系扩展建议...")
    extension_rows = build_extension_suggestions(rows, mapper_a, original_key)
    print(f"体系扩展建议数量: {len(extension_rows)}")
    llm_stats = get_llm_stats()

    print("写入结果文件...")
    append_jsonl(rows, jsonl_path)
    write_excel(rows, extension_rows, llm_stats, output_path, jsonl_path, started, args.embedder,
                strategy, args.llm_budget, args.route_b_llm, args.route_b_llm_budget,
                args.sample_seed, args.sample_mode, args.sample_size)
    config.DEEPSEEK_API_KEY = original_key
    print(f"完成: {output_path}")


if __name__ == "__main__":
    main()
