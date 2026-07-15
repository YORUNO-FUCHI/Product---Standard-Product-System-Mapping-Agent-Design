# -*- coding: utf-8 -*-
"""影响评估：旧"字面重叠一票否决"会误杀多少正确映射，新语义门槛救回多少。

复用校准脚本产出的 bge 节点矩阵缓存（cache/st_node_emb_calib.npy）。
对标注集采样（同 seed），对每条"产品→标准答案节点"：
  - old_lit_false = core_overlap(产品, 节点名) is False  → 旧门槛会判其不可靠（除非精确匹配）
  - sem = cos(产品, 节点)                                → 新门槛的语义信号
统计：
  A. 正确映射里字面不重叠(会被旧门槛误杀)的占比
  B. 其中语义 >= τ 被新门槛救回的占比
产物：cache/sem_gate_impact.json / .txt
"""
import json
import os
import sys
import time

import numpy as np
from openpyxl import load_workbook

AGENT_ROOT = r"D:\Code\finer_opd\related_paper\Product---Standard-Product-System-Mapping-Agent-Design-test01\Product---Standard-Product-System-Mapping-Agent-Design-test01\agent"
sys.path.insert(0, AGENT_ROOT)

from product_mapper import config              # noqa: E402
from product_mapper.taxonomy import load_nodes  # noqa: E402
from product_mapper.embedder import STEmbedder  # noqa: E402
from product_mapper.text import core_overlap    # noqa: E402

LABEL_XLSX = os.path.join(os.path.dirname(AGENT_ROOT), "清洗完毕_6千组_标准答案标注.xlsx")
CACHE_DIR = config.CACHE_DIR
MAT_CACHE = CACHE_DIR / "st_node_emb_calib.npy"
IDS_CACHE = CACHE_DIR / "st_node_ids_calib.npy"
CALIB_JSON = CACHE_DIR / "sem_threshold_calibration.json"
OUT_JSON = CACHE_DIR / "sem_gate_impact.json"
OUT_TXT = CACHE_DIR / "sem_gate_impact.txt"

SAMPLE_N = 1500
SEED = 20260715
COL_PRODUCT, COL_NODE_ID = 5, 6


def log(m):
    print(m, flush=True)


def load_label_rows(id_set):
    wb = load_workbook(LABEL_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        try:
            product, nid = r[COL_PRODUCT], r[COL_NODE_ID]
        except (IndexError, TypeError):
            continue
        if not product or nid is None:
            continue
        product = str(product).strip()
        try:
            nid = int(nid)
        except (ValueError, TypeError):
            continue
        if product and nid in id_set:
            rows.append((product, nid))
    wb.close()
    return rows


def main():
    tau = config.SEM_SIM_THRESHOLD
    tau_strong = tau   # 已取消 strong 单独采信路径，仅报 tau 口径
    log(f"使用阈值：SEM_SIM_THRESHOLD={tau}")

    nodes = load_nodes()
    by_id = {n.id: n for n in nodes}
    emb = STEmbedder()
    mat = np.load(MAT_CACHE)
    ids = np.load(IDS_CACHE)
    id2idx = {int(i): k for k, i in enumerate(ids)}

    rows = load_label_rows(set(id2idx.keys()))
    rng = np.random.default_rng(SEED)
    if len(rows) > SAMPLE_N:
        idx = rng.choice(len(rows), size=SAMPLE_N, replace=False)
        rows = [rows[i] for i in idx]
    log(f"采样行数：{len(rows)}，计算中…")

    n = 0
    lit_false = 0                 # 正确映射里字面不重叠（旧门槛会误杀）
    rescued_tau = 0               # 其中语义 >= tau
    rescued_strong = 0            # 其中语义 >= tau_strong
    pos_sims = []
    t0 = time.time()
    for k, (product, nid) in enumerate(rows):
        node = by_id.get(nid)
        if node is None:
            continue
        n += 1
        pv = emb.encode_one(product).astype(np.float32)
        sem = float(pv @ mat[id2idx[nid]])
        pos_sims.append(sem)
        lit = core_overlap(product, node.name)
        if lit is False:
            lit_false += 1
            if sem >= tau:
                rescued_tau += 1
            if sem >= tau_strong:
                rescued_strong += 1
        if (k + 1) % 300 == 0:
            log(f"  {k+1}/{len(rows)}  {time.time()-t0:.0f}s")

    pos = np.array(pos_sims)
    result = {
        "tau": tau,
        "tau_strong": tau_strong,
        "n": n,
        "lit_false": lit_false,
        "lit_false_ratio": round(lit_false / n, 4) if n else 0,
        "rescued_by_tau": rescued_tau,
        "rescued_by_tau_ratio_of_litfalse": round(rescued_tau / lit_false, 4) if lit_false else 0,
        "rescued_by_strong": rescued_strong,
        "rescued_by_strong_ratio_of_litfalse": round(rescued_strong / lit_false, 4) if lit_false else 0,
        "pos_sim_mean": round(float(pos.mean()), 4) if n else 0,
    }
    CACHE_DIR.mkdir(exist_ok=True)
    OUT_JSON.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")

    lines = [
        "Route B 语义门槛 — 影响评估",
        f"阈值: tau={tau}  tau_strong={tau_strong}",
        f"正确映射样本数 n = {n}",
        f"A. 字面不重叠(旧门槛会误杀)的正确映射: {lit_false}  占比 {result['lit_false_ratio']*100:.1f}%",
        f"B. 其中被语义(>=tau)救回: {rescued_tau}  占误杀的 {result['rescued_by_tau_ratio_of_litfalse']*100:.1f}%",
        f"   其中被强语义(>=tau_strong)救回: {rescued_strong}  占误杀的 {result['rescued_by_strong_ratio_of_litfalse']*100:.1f}%",
        f"正确映射的语义相似度均值: {result['pos_sim_mean']}",
    ]
    OUT_TXT.write_text("\n".join(lines), encoding="utf-8")
    for ln in lines:
        log(ln)
    log(f"[saved] {OUT_JSON}")


if __name__ == "__main__":
    main()
