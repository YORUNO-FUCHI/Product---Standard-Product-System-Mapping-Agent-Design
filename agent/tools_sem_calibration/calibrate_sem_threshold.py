# -*- coding: utf-8 -*-
"""校准 v2：判断语义门槛应该用"绝对余弦"还是"排名"。

复用已缓存的 bge 节点矩阵。对每条 (产品, 正确节点)：
  - pos            = cos(产品, 正确节点)
  - rank           = 正确节点在全部 21090 节点里按余弦的名次（0=最相似）
  - neg_random     = cos(产品, 随机若干节点) 的均值（语义无关背景水平 → 漂移大致落这）
  - neg_hard       = max_{node!=正确} cos（最容易混的错节点，最悲观）
输出：
  - 正确节点 rank 命中 top-10/30/50/100/300 的比例  → 判断"排名门槛"是否可行
  - pos / neg_random / neg_hard 分布                 → 判断"绝对阈值"能挡漂移到什么程度
产物：cache/sem_calib_v2.json / .txt
"""
import json
import os
import sys
import time

import numpy as np
from openpyxl import load_workbook

AGENT_ROOT = r"D:\Code\finer_opd\related_paper\Product---Standard-Product-System-Mapping-Agent-Design-test01\Product---Standard-Product-System-Mapping-Agent-Design-test01\agent"
sys.path.insert(0, AGENT_ROOT)

from product_mapper import config              # noqa: E402
from product_mapper.taxonomy import load_nodes  # noqa: E402
from product_mapper.embedder import STEmbedder  # noqa: E402

LABEL_XLSX = os.path.join(os.path.dirname(AGENT_ROOT), "清洗完毕_6千组_标准答案标注.xlsx")
CACHE_DIR = config.CACHE_DIR
MAT_CACHE = CACHE_DIR / "st_node_emb_calib.npy"
IDS_CACHE = CACHE_DIR / "st_node_ids_calib.npy"
OUT_JSON = CACHE_DIR / "sem_calib_v2.json"
OUT_TXT = CACHE_DIR / "sem_calib_v2.txt"

SAMPLE_N, SEED = 1500, 20260715
COL_PRODUCT, COL_NODE_ID = 5, 6
RAND_NEG_PER = 5


def log(m):
    print(m, flush=True)


def load_label_rows(id_set):
    wb = load_workbook(LABEL_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        try:
            product, nid = r[COL_PRODUCT], r[COL_NODE_ID]
        except (IndexError, TypeError):
            continue
        if not product or nid is None:
            continue
        product = str(product).strip()
        try:
            nid = int(nid)
        except (ValueError, TypeError):
            continue
        if product and nid in id_set:
            rows.append((product, nid))
    wb.close()
    return rows


def pct(a, ps):
    return {f"p{p}": round(float(np.percentile(a, p)), 4) for p in ps}


def main():
    nodes = load_nodes()
    emb = STEmbedder()
    mat = np.load(MAT_CACHE)
    ids = np.load(IDS_CACHE)
    id2idx = {int(i): k for k, i in enumerate(ids)}
    N = mat.shape[0]

    rows = load_label_rows(set(id2idx.keys()))
    rng = np.random.default_rng(SEED)
    if len(rows) > SAMPLE_N:
        idx = rng.choice(len(rows), size=SAMPLE_N, replace=False)
        rows = [rows[i] for i in idx]
    log(f"采样 {len(rows)} 行，计算 rank/pos/neg …")

    pos, ranks, neg_rand, neg_hard = [], [], [], []
    t0 = time.time()
    for k, (product, nid) in enumerate(rows):
        ci = id2idx[nid]
        pv = emb.encode_one(product).astype(np.float32)
        sims = mat @ pv
        p = float(sims[ci])
        pos.append(p)
        # rank：有多少节点余弦 >= 正确节点（越小越好，0=最相似）
        ranks.append(int((sims > p).sum()))
        rnd = rng.integers(0, N, size=RAND_NEG_PER)
        neg_rand.append(float(sims[rnd].mean()))
        sims[ci] = -1.0
        neg_hard.append(float(sims.max()))
        if (k + 1) % 300 == 0:
            log(f"  {k+1}/{len(rows)}  {time.time()-t0:.0f}s")

    pos = np.array(pos); ranks = np.array(ranks)
    neg_rand = np.array(neg_rand); neg_hard = np.array(neg_hard)

    topk = {f"top{k}": round(float((ranks < k).mean()), 4) for k in (5, 10, 30, 50, 100, 300, 1000)}

    # 绝对阈值：正确保留率 vs 随机负样本(漂移)接受率
    grid = np.round(np.arange(0.20, 0.751, 0.01), 3)
    abs_curve = []
    for tau in grid:
        tpr = float((pos >= tau).mean())
        fpr_rand = float((neg_rand >= tau).mean())
        fpr_hard = float((neg_hard >= tau).mean())
        abs_curve.append({"tau": float(tau), "tpr": round(tpr, 4),
                          "fpr_random": round(fpr_rand, 4), "fpr_hard": round(fpr_hard, 4)})

    # 选：保住正确>=90%的最高 τ（尽量挡漂移）
    keep90 = [c for c in abs_curve if c["tpr"] >= 0.90]
    tau_keep90 = max(keep90, key=lambda c: c["tau"])["tau"] if keep90 else 0.2
    row90 = next(c for c in abs_curve if c["tau"] == tau_keep90)

    result = {
        "n": len(rows),
        "rank_topk_hit": topk,
        "rank_stats": {"mean": round(float(ranks.mean()), 1), **pct(ranks, [50, 75, 90, 95])},
        "pos": {"mean": round(float(pos.mean()), 4), **pct(pos, [10, 25, 50, 90])},
        "neg_random": {"mean": round(float(neg_rand.mean()), 4), **pct(neg_rand, [50, 90, 95, 99])},
        "neg_hard": {"mean": round(float(neg_hard.mean()), 4), **pct(neg_hard, [50, 90])},
        "abs_threshold_keep_correct>=90%": {"tau": tau_keep90, **row90},
        "abs_curve": abs_curve,
    }
    CACHE_DIR.mkdir(exist_ok=True)
    OUT_JSON.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")

    lines = [
        "Route B 语义门槛校准 v2",
        f"样本 n = {len(rows)}",
        "",
        "[排名门槛可行性] 正确节点落在产品语义最近的 top-K 的比例：",
        f"  top10={topk['top10']*100:.1f}%  top30={topk['top30']*100:.1f}%  "
        f"top50={topk['top50']*100:.1f}%  top100={topk['top100']*100:.1f}%  top300={topk['top300']*100:.1f}%",
        f"  正确节点名次 mean={result['rank_stats']['mean']}  中位={result['rank_stats']['p50']}  "
        f"p90={result['rank_stats']['p90']}  (总 {mat.shape[0]} 节点)",
        "",
        "[绝对余弦分布]",
        f"  正确节点   pos:        mean={result['pos']['mean']} p10={result['pos']['p10']} p50={result['pos']['p50']} p90={result['pos']['p90']}",
        f"  随机节点(漂移背景):    mean={result['neg_random']['mean']} p90={result['neg_random']['p90']} p95={result['neg_random']['p95']} p99={result['neg_random']['p99']}",
        f"  最相似错节点(最悲观):  mean={result['neg_hard']['mean']} p90={result['neg_hard']['p90']}",
        "",
        f"[绝对阈值·保住正确>=90%] tau={row90['tau']}  "
        f"(TPR={row90['tpr']}, 误收随机漂移 FPR_random={row90['fpr_random']}, 误收最相似错节点 FPR_hard={row90['fpr_hard']})",
    ]
    OUT_TXT.write_text("\n".join(lines), encoding="utf-8")
    for ln in lines:
        log(ln)
    log(f"[saved] {OUT_JSON}")


if __name__ == "__main__":
    main()
