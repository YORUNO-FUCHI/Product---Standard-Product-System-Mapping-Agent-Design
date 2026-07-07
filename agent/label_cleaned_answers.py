"""Generate a conservative standard-answer annotation workbook for cleaned products.

This script does not overwrite the original workbook. It expands multi-product rows,
matches each product against the standard taxonomy, and writes a review-friendly
Excel workbook with node ids, standard paths, confidence levels, and review flags.
"""
from __future__ import annotations

"""
它是用来给清洗后的 6000 条数据生成“标准答案预标注表”的。
它做的事大概是：
读取 清洗完毕_6千组.xlsx
读取产品标准体系节点
如果一行里有多个产品，用 、；; 拆成多个产品
尝试给每个产品匹配一个标准体系节点
输出一个新的 Excel：
清洗完毕_6千组_标准答案标注.xlsx
但是上面的表由AI生成，所以并没有得到校验，所以仅供参考
"""

import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
PROJECT_ROOT = ROOT.parent
sys.path.insert(0, str(ROOT))

from product_mapper import config  # noqa: E402
from product_mapper.recall import MemoryRecall  # noqa: E402
from product_mapper.rerank import _fuse  # noqa: E402
from product_mapper.taxonomy import Node, load_nodes  # noqa: E402
from product_mapper.text import norm  # noqa: E402


INPUT_XLSX = PROJECT_ROOT / "清洗完毕_6千组.xlsx"
OUTPUT_XLSX = PROJECT_ROOT / "清洗完毕_6千组_标准答案标注.xlsx"

FULL_HEADERS = [
    "原sheet",
    "原序号",
    "拆分序号",
    "原始名称",
    "清洗后名称",
    "拆分后产品名",
    "标准node_id",
    "标准节点名称",
    "标准路径",
    "是否叶子节点",
    "标注状态",
    "标注把握度",
    "标注理由",
    "备注",
]

SUMMARY_HEADERS = ["指标", "数值", "说明"]
TOP_HEADERS = ["排名", "标准node_id/方向", "名称/路径", "数量"]

SPLIT_RE = re.compile(r"\s*[、；;]\s*")


@dataclass
class LabelResult:
    node: Node | None
    status: str
    confidence: str
    reason: str
    note: str = ""


def _safe_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def split_products(text: str) -> list[str]:
    """Split obvious product-list rows while keeping the original row structure traceable."""
    text = _safe_text(text)
    if not text:
        return []
    parts = [p.strip(" \t\r\n,，.。") for p in SPLIT_RE.split(text)]
    parts = [p for p in parts if p]
    return parts or [text]


def build_exact_index(nodes: Iterable[Node]) -> dict[str, Node]:
    exact: dict[str, Node] = {}
    for node in nodes:
        keys = [node.name, *node.synonyms]
        for key in keys:
            nkey = norm(key)
            if nkey and nkey not in exact:
                exact[nkey] = node
    return exact


def classify_by_scores(product: str, ordered) -> LabelResult:
    """Conservative assisted labeling from local recall scores.

    The result is intentionally cautious: weak matches become review items instead
    of being treated as final gold labels.
    """
    if not ordered:
        return LabelResult(
            node=None,
            status="体系外/待新增",
            confidence="低",
            reason="标准体系召回无候选",
            note="建议人工判断是否需要新增标准节点",
        )

    top = ordered[0]
    second = ordered[1] if len(ordered) > 1 else None
    margin = top.fused - (second.fused if second else 0.0)
    node = top.node
    score_note = (
        f"trgm={top.trgm:.3f}; vec={top.vec:.3f}; "
        f"fused={top.fused:.3f}; margin={margin:.3f}"
    )

    if not node.is_leaf:
        return LabelResult(
            node=node,
            status="待复核",
            confidence="中" if top.fused >= 1.0 or top.trgm >= 0.35 else "低",
            reason="最相近候选不是叶子节点，需人工下钻确认",
            note=score_note,
        )

    if top.trgm >= 0.25 or (top.trgm >= 0.12 and top.vec >= 0.25 and margin >= 0.20):
        return LabelResult(
            node=node,
            status="已标注",
            confidence="高",
            reason="本地召回强匹配到叶子节点",
            note=score_note,
        )

    if top.trgm >= 0.08 or top.vec >= 0.28:
        return LabelResult(
            node=node,
            status="待复核",
            confidence="中",
            reason="候选节点相关，但需要人工确认是否为最细标准节点",
            note=score_note,
        )

    if top.trgm >= 0.04 or top.vec >= 0.20:
        return LabelResult(
            node=node,
            status="待复核",
            confidence="低",
            reason="弱相关候选，仅作为人工复核参考",
            note=score_note,
        )

    return LabelResult(
        node=None,
        status="体系外/待新增",
        confidence="低",
        reason="标准体系中未找到足够相近的节点",
        note=f"最近候选：{node.path_str}；{score_note}",
    )


def label_product(product: str, exact_index: dict[str, Node], recaller: MemoryRecall) -> LabelResult:
    product = _safe_text(product)
    if not product:
        return LabelResult(
            node=None,
            status="体系外/待新增",
            confidence="低",
            reason="产品名为空",
            note="清洗后和原始名称均无法提供有效产品名",
        )

    exact = exact_index.get(norm(product))
    if exact:
        if exact.is_leaf:
            return LabelResult(
                node=exact,
                status="已标注",
                confidence="高",
                reason="产品名与标准节点名或同义词精确一致",
            )
        return LabelResult(
            node=exact,
            status="待复核",
            confidence="中",
            reason="精确命中标准节点，但该节点不是叶子节点",
            note="建议人工确认是否需要下钻到子节点",
        )

    candidates = recaller.recall(product)
    ordered = _fuse(candidates)
    return classify_by_scores(product, ordered)


def row_from_label(sheet_name: str, row, split_idx: int, product: str, label: LabelResult) -> list:
    node = label.node
    return [
        sheet_name,
        row["seq"],
        split_idx,
        row["raw"],
        row["cleaned"],
        product,
        node.id if node else "",
        node.name if node else "",
        node.path_str if node else "",
        "是" if node and node.is_leaf else ("否" if node else ""),
        label.status,
        label.confidence,
        label.reason,
        label.note,
    ]


def read_cleaned_rows(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows: list[dict] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for values in ws.iter_rows(min_row=2, values_only=True):
                if not values or all(v is None for v in values):
                    continue
                raw = _safe_text(values[1] if len(values) > 1 else "")
                cleaned = _safe_text(values[2] if len(values) > 2 else "")
                status = _safe_text(values[3] if len(values) > 3 else "")
                seq = values[0] if values and values[0] is not None else ""
                rows.append({
                    "sheet": sheet_name,
                    "seq": seq,
                    "raw": raw,
                    "cleaned": cleaned,
                    "status": status,
                })
    finally:
        wb.close()
    return rows


def append_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def autosize(ws, max_width: int = 60) -> None:
    for col_idx, cells in enumerate(ws.columns, 1):
        width = 10
        for cell in cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(max_width, len(value) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_workbook(full_rows: list[list], source_rows: int, split_rows: int) -> None:
    wb = Workbook()
    ws_full = wb.active
    ws_full.title = "完整标注结果"
    append_header(ws_full, FULL_HEADERS)
    for row in full_rows:
        ws_full.append(row)

    status_counter = Counter(r[10] for r in full_rows)
    confidence_counter = Counter(r[11] for r in full_rows)
    node_counter = Counter((r[6], r[7], r[8]) for r in full_rows if r[6])
    pending_rows = [r for r in full_rows if r[10] == "待复核"]
    outside_rows = [r for r in full_rows if r[10] == "体系外/待新增"]

    ws_pending = wb.create_sheet("待复核样本")
    append_header(ws_pending, FULL_HEADERS)
    for row in pending_rows:
        ws_pending.append(row)

    ws_outside = wb.create_sheet("体系外待新增样本")
    append_header(ws_outside, FULL_HEADERS)
    for row in outside_rows:
        ws_outside.append(row)

    ws_stats = wb.create_sheet("标注统计")
    append_header(ws_stats, SUMMARY_HEADERS)
    stats = [
        ("原始行数", source_rows, "清洗完毕_6千组.xlsx 的原始数据行数"),
        ("拆分后产品数", split_rows, "多产品行拆分后的产品记录数"),
        ("已标注数量", status_counter.get("已标注", 0), "高把握或强匹配样本"),
        ("待复核数量", status_counter.get("待复核", 0), "需人工确认的样本"),
        ("体系外/待新增数量", status_counter.get("体系外/待新增", 0), "未找到足够相近标准节点的样本"),
        ("高把握数量", confidence_counter.get("高", 0), "标注把握度=高"),
        ("中把握数量", confidence_counter.get("中", 0), "标注把握度=中"),
        ("低把握数量", confidence_counter.get("低", 0), "标注把握度=低"),
    ]
    for item in stats:
        ws_stats.append(item)

    start = ws_stats.max_row + 3
    ws_stats.cell(start, 1, "Top 20 高频标准节点").font = Font(bold=True)
    for offset, header in enumerate(TOP_HEADERS, 0):
        ws_stats.cell(start + 1, 1 + offset, header).font = Font(bold=True)
    for rank, ((node_id, name, path), count) in enumerate(node_counter.most_common(20), 1):
        ws_stats.append([rank, node_id, f"{name} | {path}", count])

    start = ws_stats.max_row + 3
    ws_stats.cell(start, 1, "Top 20 待新增方向").font = Font(bold=True)
    for offset, header in enumerate(["排名", "建议方向", "示例/说明", "数量"], 0):
        ws_stats.cell(start + 1, 1 + offset, header).font = Font(bold=True)
    direction_counter = Counter()
    for row in outside_rows:
        note = row[13] or row[12] or "未识别方向"
        direction = note.split("；", 1)[0]
        direction_counter[direction] += 1
    for rank, (direction, count) in enumerate(direction_counter.most_common(20), 1):
        ws_stats.append([rank, direction, "", count])

    for ws in wb.worksheets:
        autosize(ws)

    wb.save(OUTPUT_XLSX)


def main() -> None:
    if not INPUT_XLSX.exists():
        raise FileNotFoundError(f"Input workbook not found: {INPUT_XLSX}")

    # Keep this workbook generation local and deterministic.
    config.DEEPSEEK_API_KEY = ""
    config.RECALL_BACKEND = "memory"
    config.EMBEDDER = "hash"

    print("加载标准产品体系...")
    nodes = load_nodes()
    exact_index = build_exact_index(nodes)
    recaller = MemoryRecall(nodes, embedder_type="hash")

    print("读取清洗数据...")
    source_rows = read_cleaned_rows(INPUT_XLSX)
    full_rows: list[list] = []
    product_to_label: dict[str, LabelResult] = {}
    consistency: dict[str, set[int]] = defaultdict(set)

    print("开始生成预标注...")
    for idx, row in enumerate(source_rows, 1):
        base_text = row["cleaned"] or row["raw"]
        products = split_products(base_text)
        for split_idx, product in enumerate(products, 1):
            label = product_to_label.get(product)
            if label is None:
                label = label_product(product, exact_index, recaller)
                product_to_label[product] = label
            if label.node:
                consistency[product].add(label.node.id)
            full_rows.append(row_from_label(row["sheet"], row, split_idx, product, label))
        if idx % 500 == 0:
            print(f"  已处理 {idx}/{len(source_rows)} 原始行")

    split_rows = len(full_rows)
    print("写入 Excel...")
    write_workbook(full_rows, source_rows=len(source_rows), split_rows=split_rows)
    print(f"完成: {OUTPUT_XLSX}")
    print(f"原始行数: {len(source_rows)} | 拆分后产品数: {split_rows}")


if __name__ == "__main__":
    main()
